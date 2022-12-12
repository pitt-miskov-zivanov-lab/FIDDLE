# from __future__ import division

import os
import re
# import copy
import pickle
import random
# import openpyxl
import networkx as nx
# from networkx import *
# import multiprocessing
import matplotlib.pyplot as plt
# from collections import Counter
import Simulator.simulator as sim
from datetime import datetime as dt
# from joblib import Parallel, delayed
from openpyxl import Workbook, load_workbook
# from networkx.drawing.nx_agraph import graphviz_layout

# import time
# import os, sys
# from scoring import *
# from DFS_endValueLoop import *
# from BFS_endValueLoop import *
# from extension_util import extend_model

# import argparse
# import BDFA_scoring as scoring
# import BDFA_extension_util as eu
# from collections import defaultdict
# from joblib import Parallel, delayed


def add_regulation_to_directed_network(G,positive_probability):
    """Adds positive and negative regulation attributes to each edges of graph G.

    Parameters
    ----------
    G : nx.Graph()
        A randomly created graph using one of networkx's many random graph generators. 
        
    positive_probability : float [0-1]
        The probability that a edge in the network is positive.


    Returns
    -------
    H : nx.DiGraph()
        This function returns a directed graph with a regulation attribute for each edge based on the positive_probability specified. 
    """    
    #Initialize the output graph
    H = nx.DiGraph(weight=1,typ='+')

    #Seed the random number generator
    random.seed()

    #Iterate through graph G, assigning regulation before adding it to H. 
    for edge in G.edges():

        #Identify source and target
        from_node = str(edge[0])
        to_node = str(edge[1])

        #Select a random number
        chance = random.random()
        
        #Using a random number to determine if the regulation is positive or negative depending upon the 'positive_probability'.
        if chance > positive_probability: H.add_edge(from_node,to_node,weight=1,typ='-')
        else: H.add_edge(from_node,to_node,weight=1,typ='+')

    return H


def add_direction_to_undirected_network(G):
    """Adds directionality to an undirected graph G.

    Parameters
    ----------
    G : nx.Graph()
        A randomly created graph using one of networkx's many random graph generators. 


    Returns
    -------
    H : nx.DiGraph()
        This function returns a directed graph. 
    """  
    #Initialize the output graph
    H = nx.DiGraph(weight=1,typ='+')

    #Iterate through graph G, assigning regulation before adding it to H. 
    for edge in G.edges():

        #Identify source and target
        from_node = str(edge[0])
        to_node = str(edge[1])

        #Create matching edge in the DiGraph... now with directionality!
        H.add_edge(from_node,to_node,weight=1)

    return H


def graph_maker(network_type, nodes, positive_probability, edge_probability=0.5, attaching_edges=2):
    """Automatically creates a directed network with positive and negative edges of a particular type.

    Parameters
    ----------
    network_type : int
        A quick and easy manner to specify the type of graph you would like created:
            1 = networkx.gn_graph() = a growing network (GN) digraph with n nodes.
            2 = networkx.gnr_graph() = a growing network with redirection (GNR) digraph with n nodes and redirection probability p.
            3 = networkx.gnc_graph() = a growing network with copying (GNC) digraph with n nodes.
            4 = networkx.gnp_random_graph() = an Erdős-Rényi graph or a binomial graph.
            5 = networkx.barabasi_albert_graph() = a random graph according to the Barabási–Albert preferential attachment model.
            More information is available at: https://networkx.org/documentation/stable/reference/generators.html

    nodes : int
        The desired number of nodes in the network.

    positive_probability : float [0-1]
        The probability that a edge in the network is positive.

    edge_probability : float [0-1] #TODO
        A network-type parameter which changes based on the type of network selected. Only affects type 2 and 4.
        In network_type == 2: 'edge_probability' referes the the probability and edge is redirected.
        In network_type == 4: 'edge_probability' referes the the probability that an edge is created.

    attaching_edges : int [must be >=1 but <nodes]
        Only affectes network_type == 5. This parameter specifies the number of edges to attach from a new node to existing nodes.


    Returns
    -------
    Network : nx.DiGraph()
        This function returns a directed network with positive and negative edges based on the input parameters specified. If the network cannot be created based on your specifications, \'Network\' is returned as \'None\'. 

    """
    #Initialize Network Variable
    Network = None

    #Creation of a growing network (GN) digraph with n nodes.
    if network_type == 1:
        G = nx.gn_graph(nodes)

    #Creation of a growing network with redirection (GNR) digraph with n nodes and redirection probability p.
    elif network_type == 2:
        #Here 'edge_probability' referes the the probability and edge is redirected.
        G = nx.gnr_graph(nodes, edge_probability)

    #Creation of a growing network with copying (GNC) digraph with n nodes.
    elif network_type == 3:
        G = nx.gnc_graph(nodes)

    #Creation of an Erdős-Rényi graph or a binomial graph.
    elif network_type == 4:
        #Here 'edge_probability' referes the the probability that an edge is created.
        G = nx.gnp_random_graph(nodes, edge_probability, directed = True)
    
    #Creation of a random graph according to the Barabási–Albert preferential attachment model.
    elif network_type == 5:
        # In this type of graph, we must:
        #       specify the initial number of starting edges (attaching_edges)
        #       transform the undirected graph into a DiGraph()
        G = nx.barabasi_albert_graph(nodes, attaching_edges)
        G = add_direction_to_undirected_network(G)

    #Specified network_type was invalid
    else:
        print("Variable \'network_type\' not a recognized input.")
        return Network

    Network = add_regulation_to_directed_network(G,positive_probability)
    return Network


def get_model_edges(G):
    """Returns a list of all of G's edges.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 


    Returns
    -------
    list_of_edges : list of lists
        A list of lists with all of the edges in G. Each edge is represented as a list of [source, target, regulation]. 
        For example 1 -> 2 would be represented as [1,2,"+"].
        For example 1 -| 2 would be represented as [1,2,"-"].
    """ 
    #Initialize the return variable     
    list_of_edges = []

    #Iterate through the edges in G
    for edge in G.edges():

        #Identify the source and target nodes
        from_node = edge[0]
        to_node = edge[1]

        #Identify the regulation type as an attribute saved in 'typ'
        regulation = G[edge[0]][edge[1]]['typ']

        #Build and append the list representation of the graph edge
        build = [from_node,to_node,regulation]
        list_of_edges.append(build)
        
    return list_of_edges


def return_model_stats(G, network_type, nodes, positive_probability, edge_probability=0.5, attaching_edges=2, verbose = True, save = False, filename = 'model_stats.txt'):
    """Returns descriptive statistics of the network. Follows similar input scheme as graph_maker.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    network_type : int
        A quick and easy manner to specify the type of graph you would like created:
            1 = networkx.gn_graph() = a growing network (GN) digraph with n nodes.
            2 = networkx.gnr_graph() = a growing network with redirection (GNR) digraph with n nodes and redirection probability p.
            3 = networkx.gnc_graph() = a growing network with copying (GNC) digraph with n nodes.
            4 = networkx.gnp_random_graph() = an Erdős-Rényi graph or a binomial graph.
            5 = networkx.barabasi_albert_graph() = a random graph according to the Barabási–Albert preferential attachment model.
            More information is available at: https://networkx.org/documentation/stable/reference/generators.html

    nodes : int
        The desired number of nodes in the network.

    positive_probability : float [0-1]
        The probability that a edge in the network is positive.

    edge_probability : float [0-1] #TODO
        A network-type parameter which changes based on the type of network selected. Only affects type 2 and 4.
        In network_type == 2: 'edge_probability' referes the the probability and edge is redirected.
        In network_type == 4: 'edge_probability' referes the the probability that an edge is created.

    attaching_edges : int [must be >=1 but <nodes]
        Only affectes network_type == 5. This parameter specifies the number of edges to attach from a new node to existing nodes.

    verbose : Bool [default = True]
        Whether or not to return network's descriptive statistics to the console. 
        True = Yes
        False = No

    save : Bool [default = False]
        Whether or not to save the network's descriptive statistics in a file. 
        True = Yes
        False = No
    
    filename : string [default = 'model_stats.txt']
        If save == True, you can specify a filename for the output to be save to. 

    """ 
    #Initialize key descriptive statistics variables:
    actual_nodes = 0
    directed_edges = 0
    undirected_edges = 0
    positive_edges = 0
    negative_edges = 0
    
    undirected_G = G.to_undirected()
    edge_list = get_model_edges(G)

    #Use networkx to calculate key statistics
    degree = nx.degree(G)
    density = nx.density(G)

    #Calculate the actual number of nodes and edges in G (as it might differ based on network creation methods)
    for node in degree:
    	actual_nodes += 1
    	undirected_edges += node[1]

    #Calculate the average node degree
    ave_degree = undirected_edges/actual_nodes

    #Count the number of positive and negative edges
    for edge in edge_list:
        directed_edges += 1
        regulation = edge[2]
        if regulation == "+": positive_edges += 1
        if regulation == "-": negative_edges += 1

    #Just some sanity checks
    if directed_edges * 2 != undirected_edges:
        print("There has been a mistake in the counting of edges!")
    if directed_edges != (positive_edges + negative_edges):
        print("There has been a mistake in the counting of edges!")
    
    #Use networkx to calculate more difficult key statistics of undirected G
    try:
        diameter = nx.diameter(G.to_undirected())
    except Exception as e:
        diameter = "Error: {error}".format(error = e)

    try:
        clustering = nx.average_clustering(G.to_undirected())
    except Exception as e:
        clustering = "Error: {error}".format(error = e)

    #Ouput key statistics if verbose == True
    if verbose:
        print("------------------ Specified Network Characteristics ------------------")
        print("Network Type: {}".format(network_type))
        print("Specified Nodes: {}".format(nodes))
        print("Specified Probability of Positive Edges: {}".format(positive_probability))
        print("Specified \'edge_probability\' (only affects Type 2 and 4): {}".format(edge_probability))
        print("Specified \'attaching_edges\' (only affects Type 5): {}".format(attaching_edges))
        print("")
        print("-------------------- Actual Network Characteristics -------------------")
        print("Number of Actual Nodes: {}".format(actual_nodes))
        print("Number of Directed Edges: {}".format(directed_edges))
        print("Number of Positive Edges: {}".format(positive_edges))
        print("Number of Negative Edges: {}".format(negative_edges))
        print("Actual Probability of Positive Edges: {}".format(positive_edges/directed_edges))
        print("")
        print("----------------- Descriptive Network Characteristics -----------------")
        print("Network Density: {}".format(density))
        print("Undirected Edges: {}".format(undirected_edges))
        print("Network Average Degree: {}".format(ave_degree))
        print("Undirected Model Diameter: {}".format(diameter))
        print("Undirected Model Clustering: {}".format(clustering))
        print("")
        print("---------------------------- List of Edges ---------------------------")
        for edge in edge_list:
            print(str(edge))


    #Save key statistics if save == True
    if save:
        try:
            with open(filename,"w") as myfile:
                myfile.write("------------------ Specified Network Characteristics ------------------"+ '\n')
                myfile.write("Network Type: {}".format(network_type)+ '\n')
                myfile.write("Specified Nodes: {}".format(nodes)+ '\n')
                myfile.write("Specified Probability of Positive Edges: {}".format(positive_probability)+ '\n')
                myfile.write("Specified \'edge_probability\' (only affects Type 2 and 4): {}".format(edge_probability)+ '\n')
                myfile.write("Specified \'attaching_edges\' (only affects Type 5): {}".format(attaching_edges)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("-------------------- Actual Network Characteristics -------------------"+ '\n')
                myfile.write("Number of Actual Nodes: {}".format(actual_nodes)+ '\n')
                myfile.write("Number of Directed Edges: {}".format(directed_edges)+ '\n')
                myfile.write("Number of Positive Edges: {}".format(positive_edges)+ '\n')
                myfile.write("Number of Negative Edges: {}".format(negative_edges)+ '\n')
                myfile.write("Actual Probability of Positive Edges: {}".format(positive_edges/directed_edges)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("----------------- Descriptive Network Characteristics -----------------"+ '\n')
                myfile.write("Network Density: {}".format(density)+ '\n')
                myfile.write("Undirected Edges: {}".format(undirected_edges)+ '\n')
                myfile.write("Network Average Degree: {}".format(ave_degree)+ '\n')
                myfile.write("Undirected Model Diameter: {}".format(diameter)+ '\n')
                myfile.write("Undirected Model Clustering: {}".format(clustering)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("---------------------------- List of Edges ---------------------------"+ '\n')
                for edge in edge_list:
                    myfile.write(str(edge)+ '\n')
                myfile.write('\n\n')
                myfile.close()

        except Exception as e:
            print(e)


def plot_network(G, save = False, filename = "test.png"):
    """Plots Network G in a quick and easy function.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    save : Bool [default = False]
        Whether or not to save the plot of the network to a file. 
        True = Yes
        False = No
    
    filename : string [default = 'test.png']
        If save == True, you can specify a filename for the output to be save to. 

    """ 
    #Standardize the figure size
    plt.figure(figsize=(5,5), dpi = 100)
    #Try to render the figure in the nice, clear planar view.
    try:
        nx.draw(G, pos=nx.planar_layout(G), node_size=600, node_color='#F62217', with_labels=True, font_weight='bold')

    #If the network is not planar, default to the spring_layout
    except Exception as e:
        nx.draw(G, pos=nx.spring_layout(G), node_size=600, node_color='#F62217', with_labels=True, font_weight='bold')

    if save:
        plt.savefig(filename)
        plt.clf()


def create_initial_values(G, initial=1):
    """Returns a dictionary with SET initial values for each node in G.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    initial : integer [default = 1]
        An integer value to assign as the initial value for all of the nodes in the network. 


    Returns
    -------
    initial_val_dic : dictionary
        A dictionary with keys for each node in G, and values for the initial value for the simulator.
    """ 
    #Initialize a dictionary to store the initail values
    initial_val_dic = {}

    #Iterate through the network edges in G
    for edge in G.edges():
        source = edge[0]
        target = edge[1]

        #Make sure we have not already initialized the source
        if source not in initial_val_dic:
            initial_val_dic[source] = initial

         #Make sure we have not already initialized the target
        if target not in initial_val_dic:
            initial_val_dic[target] = initial

    return initial_val_dic


def create_initial_values_random(G, lower_bound = 0, upper_bound = 2):
    """Returns a dictionary with RANDOM initial values for each node in G.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    lower_bound : integer [default = 0]
        An integer value which serves as the lowest possible initial value permitted for a node in the network. 

    upper_bound : integer [default = 2]
        An integer value which serves as the highest possible initial value permitted for a node in the network. 


    Returns
    -------
    initial_val_dic : dictionary
        A dictionary with keys for each node in G, and values for the initial value for the simulator.
    """ 
    #Initialize a dictionary to store the initail values
    initial_val_dic = {}

    #Seed the random generator
    random.seed()

    #Iterate through the network edges in G
    for edge in G.edges():
        source = edge[0]
        target = edge[1]

        #Make sure we have not already initialized the source
        if source not in initial_val_dic:
            initial_val_dic[source] = random.randrange(lower_bound,upper_bound+1,1)

         #Make sure we have not already initialized the target
        if target not in initial_val_dic:
            initial_val_dic[target] = random.randrange(lower_bound,upper_bound+1,1)

    return initial_val_dic


def network_to_excel(G, filename = "network_full.xlsx"):
    """Saves the Network G to an excel file in a standardized format.

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 
    
    filename : string [default = 'network_full.xlsx']
        The filename for the output to be save to.  

    """ 
    #Initialize dictionary to remember the row for each unique network node.
    name_to_row = {}

    #Initialize the excel workbook so that we can begin to save the network.
    wb = Workbook()
    ws = wb.active

    #Create standardized column headers for easy reading
    ws.cell(row=1,column=1,value='Element')
    ws.cell(row=1,column=2,value='Positive Regulators')
    ws.cell(row=1,column=3,value='Negative Regulators')
    ws.cell(row=1,column=6,value='Initial Values')
    ws.cell(row=1,column=7,value='Expected Values')

    #Initialize and begin iteration scheme to translate nodes to rows
    curr_row = 2
    for edge in G.edges():
        source = edge[0]
        target = edge[1]

        #Check to see if the source node has already been added to the excel
        if source not in name_to_row:
            #Add source to location dictionary
            name_to_row[source] = curr_row

            #Add source to excel and initialize it      
            ws.cell(row=curr_row,column=1,value=source)
            
            #Increment the current row, so we don't write over it
            curr_row += 1

        #Check to see if the target node has already been added to the excel
        if target not in name_to_row:
            #Add target to location dictionary
            name_to_row[target] = curr_row

            #Add target to excel and initialize it         
            ws.cell(row=curr_row,column=1,value=target)
            
            #Increment the current row, so we don't write over it
            curr_row += 1

        #Add the regulation information
        #Determine the type of regulation
        if G[source][target]['typ'] == '+': col = 2
        else: col = 3

        #Determining if we must append information, or if this is the first regulator of the target
        original = ws.cell(row=name_to_row[target],column=col).value
        if original == None: regulation_info = str(source)
        else: regulation_info = original + ',' + str(source)

        #Adding the regulation information into the excel
        ws.cell(row=name_to_row[target],column=col,value=regulation_info)
    
    #Saving the file
    wb.save(filename)


def update_network_file(filename, update_column, update_values):
    """Function to quickly input initial and final values for simulations. 

    Parameters
    ----------
    filename : string 
        The filename of the network excel file.  

    update_column : integer
        The column location you wish to update. Frequently this is the initial values column (6), or the final values column (7). 

    update_values : dictionary
        A dictionary with keys for each node in the network file, and the corresponding values you wish to input for that row (values).
    
    """
    #Initialize key variable to ensure not too many values were given to update
    updated_nodes = []

    #Loading the network file excel
    wb = load_workbook(filename)
    ws = wb.active

    #Initializing and iterating through the file row by row
    current_row = 2
    for idx, rows in enumerate(ws.iter_rows(),start=1):
        node = rows[0].value
        
        #Check to make sure we are in a valid row
        if (node is None) or (node.lower()=="element"): continue

        #Check to make sure we have an update value for this node
        if node in update_values.keys(): 
            #Update the row value with the correct value from the dictionary
            ws.cell(row=current_row, column=update_column, value=update_values[node])
            updated_nodes.append(node)

        else:
            print("Node: " + str(node) + " was given no value to update!")
        
        #Updating the current row
        current_row += 1

    #Check to make sure no update_values were not used:
    for key in update_values.keys(): 
        if key not in updated_nodes:
            print("Node: " + str(key) + " does not exist in this network and therefore could not be updated!") 
    #Save updated network file
    wb.save(filename)


def simulate_network(filename, simulation_runs, simulation_length, output_file = "network_trace.txt"):
    """Simulate an initialized network file. 

    Parameters
    ----------
    filename : string 
        The name of the excel file with the initialized network. 
        Example = "network_full.xlsx"

    simulation_runs : integer
        The number of simulations to run. This is necessary as the simulator is stochastic, and you can observe different behavior with each simulation. 
        Example = 10 

    simulation_length : integer
        The number of time-steps or simulation updates performed during a simulation run.
        Example = 100

    output_file : string [default = "network_trace.txt"]
        The name of the file in which the simulation trace will be saved. 

    """
    #Initialization of key variables relevant to the simulator
    output_mode = 3
    update_scheme = "ra"
    column_with_initial_values = 6

    #Load the Network into the simulator
    model_run = sim.Manager(filename, column_with_initial_values)

    #Run the simulator
    model_run.run_simulation(update_scheme, simulation_runs, simulation_length, output_file, outMode=output_mode)


def get_simulation_end_values(filename, simulation_runs):
    """Returns the end values of the network simulation, based on the simulation trace file. 
    The variable 'simulation_runs' is necessary to normalize the end_values based on the number of simulations performed. 

    Parameters
    ----------
    filename : string 
        The name of the txt file with the simulation trace output. 
        Example = "network_trace.txt"

    simulation_runs : integer
        The number of simulations to run. This is necessary as the simulator is stochastic, and you can observe different behavior with each simulation. 
        Example = 10 

    Returns
    -------
    end_values : dictionary
        A dictionary with keys for each node in the network file, and values for the node's end value after simulation.
    """
    #Initialization of key variables to hold values
    end_values = {}

    #Open the trace file, read it, parse it, and get the end values for each node
    with open(filename) as trace_file:
        trace_file.readline()
        for line in trace_file:

            #Skip the header
            if "Frequency" in line: continue

            #Cleaning the line
            line = line.strip()
            values = re.split(' ', line)

            #Parsing the node
            clutter_node = values[0]
            clean_node = clutter_node.split("|")[0]

            #Parsing the end value
            end_value = float(values[-1])
            normalized_end_value = end_value/simulation_runs

            #Saving the node and its end value
            end_values[clean_node] = normalized_end_value

    return end_values


def get_model_expected_values(filename):
    """Function to quickly retrieve the expected values for a model's simulation. 
    Used to be "get_golden"
    Parameters
    ----------
    filename : string 
        The filename of the network excel file.  

    Returns
    -------
    end_values : dictionary
        A dictionary with keys for each node in the network file, and values for the node's end value after simulation.

    """
    #Define expected column of end values
    end_val_col = 6

    #Load excel model file
    wb = load_workbook(filename)
    ws = wb.active

    #Initialize variable dictionary to hold end values
    end_values = dict()

    #Iterate through the model file to find end values
    for row in ws.iter_rows():
        #Basic filtration to ensure the wrong keys are not taken
        if row[0].value is None: continue
        if row[0].value.lower()=="variable name": continue
        if row[0].value.lower()=="element": continue
        if len(row)<end_val_col: continue
        if row[end_val_col].value is None: continue

        #Adding elements to the dictionary with their end values as floats
        end_values[row[0].value] = float(row[end_val_col].value) 

    return end_values


def get_score_between_expected_and_actual(expected_values, simulation_end_values):
    """Function to calculate the difference between the expected values for a perfect model's \
    simulation and the actual values on an unextended model.  
    
    Used to be "get_diff"

    Parameters
    ----------
    expected_values : dictionary 
        The dictionary containing the expected end values for each element in the model.  
    
    simulation_end_values : dictionary 
        The dictionary containing a simulation's end values for each element in the model.  


    Returns
    -------
    total_difference : float
        The absoluate difference between the expected values and the simulation end values as \
            a rough calculation for how different a model is from expectation.

    """
    total_difference = 0
    for key in expected_values.keys():
        try:
            total_difference += abs(expected_values[key] - simulation_end_values[key])
        except Exception as e:
            print(e)
    
    return total_difference


def find_element_locations(filename):
    """Function to locate each model element in the model file. 
    Used to be "getRow"
    
    Parameters
    ----------
    filename : string 
        The filename of the network excel file.  

    Returns
    -------
    row_location_of_element : dictionary
        A dictionary with keys for each node in the network file, and values for the node's row index.

    """
    #Initialize key dictionary
    row_location_of_element = dict()

    #Load model file and begin iterating through the rows
    wb = load_workbook(filename)
    ws = wb.active
    for idx, rows in enumerate(ws.iter_rows(), start=1):
        if rows[0].value is None or rows[0].value.lower()=="variable name": continue

        #Add element location (idx) into dictionary using element name as key
        row_location_of_element[rows[0].value] = idx

    return row_location_of_element


def extend_model_file(filename,edge_group,output_name):
    """Function to to extend a model file with a list of edges and save it as a new model file. 
    Used to be "extend_model"

    Parameters
    ----------
    filename : string 
        The filename of the network excel file.  
    
    edge_group : list
        List of edges with which to extend the model. 
        Example:
        edge_group = [0, ['1','2','+'], ['2','3','-']]
        edge_group[0] is an artifact of edge group creation. Typically used to identify the extension group. 

    output_name : string 
        The filename for the new extended network excel file.  

    """
    # #Copy a version of the most recent model file to a new location so we can alter it without affecting past work
    # os.system('cp '+current_model_file+' '+extension_model_file)

    #Identify index/row location where model elements are in the extension model file
    row_location_of_element = find_element_locations(filename)

    #If an extension includes a model element that is not yet in the model, 
    # we need to know where the first empty row is to add it in without copying over other nodes
    max_row = len(row_location_of_element)+2

    #Load model file 
    wb = load_workbook(filename)
    ws = wb.active

    #Iterate through the edges you want to add and begin adding unfamiliar nodes
    for e in edge_group[1:]:
        #If the regulator node is not already in the model file, we must add it as a new row
        if e[0] not in row_location_of_element:
            ws.cell(row=max_row,column=1,value=e[0])
            ws.cell(row=max_row,column=6,value=1) # This explicitely initializes any new nodes at 1
            row_location_of_element[e[0]] = max_row
            max_row += 1
        
        #If the regulated node is not already in the model file, we must add it as a new row
        if e[1] not in row_location_of_element:
            ws.cell(row=max_row,column=1,value=e[1])
            ws.cell(row=max_row,column=6,value=1) # This explicitely initializes any new nodes at 1
            row_location_of_element[e[1]] = max_row
            max_row += 1
        
        #Determine if the interaction is positive or negative, and set the column accordingly
        col = 2 if e[2]=='+' else 3

        #Seamlessly add the new regulator to the existing regulators without loss of information
        original = ws.cell(row=row_location_of_element[e[1]],column=col).value
        original = original + ',' if original != None else ''
        ws.cell(row=row_location_of_element[e[1]],column=col,value=original+e[0])
    
    #Save and close the extended model file
    wb.save(output_name)


def extension_network_to_excel(filename, expected_end_values):
    """Saves the Network G to an excel file in a standardized format.

    Parameters
    ----------
    filename : string 
        The filename of the network excel file.  

    expected_end_values : dictionary
        A dictionary with keys for each node in the true network file, and values for the node's end value after simulation. 

    """ 
    #Initialize key variables
    max_row = 2
    seen_nodes = []
    col_initial_values = 6
    col_expected_values = 7

    #Loading the network file excel and begin iterating through rows
    wb = load_workbook(filename)
    ws = wb.active
    for idx, rows in enumerate(ws.iter_rows(),start=1):
        #Grab node name
        node = rows[0].value

        #Skip if header
        if (node.lower()=="element"): continue

        #Otherwise, make sure row is parameterized correctly
        #Add node to list of seen nodes
        seen_nodes.append(node)
        ws.cell(row=idx, column=col_initial_values, value=1)
        ws.cell(row=idx, column=col_expected_values, value=expected_end_values[node])
        
        #Increment max_row
        max_row = idx + 1

    #Now, going back and adding/initializing nodes which should be in the model
    for node_check in expected_end_values:
        if node_check not in seen_nodes:
            ws.cell(row=max_row, column=1, value=str(node_check))
            ws.cell(row=max_row, column=col_initial_values, value=1)
            ws.cell(row=max_row, column=col_expected_values, value=expected_end_values[node_check])
            seen_nodes.append(node_check)
            max_row += 1

    wb.save(filename)


def return_extension_stats(base_name, G, real_edges, fake_edges, verbose = True, save = False, filename = 'model_stats.txt'):
    """Returns descriptive statistics of the network for extension. 

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    real_edges : list
        A list of edges that were removed from the baseline model.

    fake_edges : list
        A list of edges that were never in the baseline model but used to test extension fidelity.

    verbose : Bool [default = True]
        Whether or not to return network's descriptive statistics to the console. 
        True = Yes
        False = No

    save : Bool [default = False]
        Whether or not to save the network's descriptive statistics in a file. 
        True = Yes
        False = No
    
    filename : string [default = 'model_stats.txt']
        If save == True, you can specify a filename for the output to be save to. 

    """ 
    #Initialize key descriptive statistics variables:
    actual_nodes = 0
    directed_edges = 0
    undirected_edges = 0
    positive_edges = 0
    negative_edges = 0
    
    undirected_G = G.to_undirected()
    edge_list = get_model_edges(G)

    #Use networkx to calculate key statistics
    degree = nx.degree(G)
    density = nx.density(G)

    #Calculate the actual number of nodes and edges in G (as it might differ based on network creation methods)
    for node in degree:
    	actual_nodes += 1
    	undirected_edges += node[1]

    #Calculate the average node degree
    ave_degree = undirected_edges/actual_nodes

    #Count the number of positive and negative edges
    for edge in edge_list:
        directed_edges += 1
        regulation = edge[2]
        if regulation == "+": positive_edges += 1
        if regulation == "-": negative_edges += 1

    #Just some sanity checks
    if directed_edges * 2 != undirected_edges:
        print("There has been a mistake in the counting of edges!")
    if directed_edges != (positive_edges + negative_edges):
        print("There has been a mistake in the counting of edges!")
    
    #Use networkx to calculate more difficult key statistics of undirected G
    try:
        diameter = nx.diameter(G.to_undirected())
    except Exception as e:
        diameter = "Error: {error}".format(error = e)

    try:
        clustering = nx.average_clustering(G.to_undirected())
    except Exception as e:
        clustering = "Error: {error}".format(error = e)

    #Ouput key statistics if verbose == True
    if verbose:
        print("------------------ Specified Network Characteristics ------------------")
        print("Network Name: {}".format(base_name))
        print("")
        print("-------------------- Actual Network Characteristics -------------------")
        print("Number of Actual Nodes: {}".format(actual_nodes))
        print("Number of Directed Edges: {}".format(directed_edges))
        print("Number of Positive Edges: {}".format(positive_edges))
        print("Number of Negative Edges: {}".format(negative_edges))
        print("Actual Probability of Positive Edges: {}".format(positive_edges/directed_edges))
        print("")
        print("----------------- Descriptive Network Characteristics -----------------")
        print("Network Density: {}".format(density))
        print("Undirected Edges: {}".format(undirected_edges))
        print("Network Average Degree: {}".format(ave_degree))
        print("Undirected Model Diameter: {}".format(diameter))
        print("Undirected Model Clustering: {}".format(clustering))
        print("")
        print("---------------------------- List of Edges ---------------------------")
        for edge in edge_list:
            print(str(edge))
        print("")
        print("----------------------- List of Real Extensions ----------------------")
        for edge in real_edges:
            print(str(edge))
        print("")
        print("----------------------- List of Fake Extensions ----------------------")
        for edge in fake_edges:
            print(str(edge))

    #Save key statistics if save == True
    if save:
        try:
            with open(filename,"w") as myfile:
                myfile.write("------------------ Specified Network Characteristics ------------------"+ '\n')
                myfile.write("Network Name: {}".format(base_name))
                myfile.write(""+ '\n')
                myfile.write("-------------------- Actual Network Characteristics -------------------"+ '\n')
                myfile.write("Number of Actual Nodes: {}".format(actual_nodes)+ '\n')
                myfile.write("Number of Directed Edges: {}".format(directed_edges)+ '\n')
                myfile.write("Number of Positive Edges: {}".format(positive_edges)+ '\n')
                myfile.write("Number of Negative Edges: {}".format(negative_edges)+ '\n')
                myfile.write("Actual Probability of Positive Edges: {}".format(positive_edges/directed_edges)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("----------------- Descriptive Network Characteristics -----------------"+ '\n')
                myfile.write("Network Density: {}".format(density)+ '\n')
                myfile.write("Undirected Edges: {}".format(undirected_edges)+ '\n')
                myfile.write("Network Average Degree: {}".format(ave_degree)+ '\n')
                myfile.write("Undirected Model Diameter: {}".format(diameter)+ '\n')
                myfile.write("Undirected Model Clustering: {}".format(clustering)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("---------------------------- List of Edges ---------------------------"+ '\n')
                for edge in edge_list:
                    myfile.write(str(edge)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("----------------------- List of Real Extensions ----------------------"+ '\n')
                for edge in real_edges:
                    myfile.write(str(edge)+ '\n')
                myfile.write(""+ '\n')
                myfile.write("----------------------- List of Fake Extensions ----------------------"+ '\n')
                for edge in fake_edges:
                    myfile.write(str(edge)+ '\n')
                myfile.write('\n\n')
                myfile.close()

        except Exception as e:
            print(e)


def create_extendable_models(G, output_folder, removal_probabilities, expected_end_values, model_name = 'network', positive_probability = 0.5):
    """Function to create identical copies of the graph G, with different probabilities of missing edges. This is done in one function to \
        ensure that an edges missing in a copy of G at low probability is also missing from a copy of G at a higher probability. This allows \
        us to control for some edges being more "important" to model performance than others.  

    Parameters
    ----------
    G : nx.DiGraph()
        A randomly created graph using one of networkx's many random graph generators. 

    output_folder : file path (string)
        A relative or absolute filepath that will house the new models 

    removal_probabilities : list
        A list of probabilities (floats) in ascending order of edge removal
        Example: removal_probabilities = [0.1, 0.5] -> to have two models created missing about 10% and 50% of their edges.
        BE AWARE: removal_probabilities = [0.5, 0.1] -> WRONG... not in ascending order
        #TODO maybe I should just sort the list myself.

    expected_end_values : dictionary
        A dictionary with keys for each node in the true network file, and values for the node's end value after simulation. 

    model_name : string [default = 'network']
        Optional name to serve as base of created files. 

    positive_probability : float [0-1]
        The probability that a newly created fake edge in the network is positive.
        This parallels the notation of network creation functions.


    Returns
    -------
    end_values : dictionary
        A dictionary with keys for each node in the network file, and values for the node's end value after simulation.

    """
    #Seed the random generator 
    random.seed()

    #Initialize key variables
    edges_in_G = list(G.edges)
    edges_not_in_G = list(nx.non_edges(G))
    edges_in_G_with_probabilities = []

    #Calculate the probability that a particular edge should be removed (we only want to do this once to not iteratively subject edges to removal repeatedly)
    for edge in edges_in_G:
        #Pick a random number
        prob = random.random()
        edges_in_G_with_probabilities.append([prob, edge])

    #Initialize ID variables
    extension_ID = 0
    fake_edge_ID = len(edges_in_G)

    #Initialize empty variables to hold extensions
    real_edges_to_add_back = []
    fake_edges_to_add_back = []

    #Defining a previous probability to not remove the same edge twice
    previous_prob = 0

    for remove_prob in removal_probabilities:

        #Remove edges in G based on probabilities and removal cutoffs
        for edge_probability, *hold_edge in edges_in_G_with_probabilities:

            edge = hold_edge[0]

            #If the random number is less than the removal probability, and not removed yet, we must remove the edge
            if (edge_probability > previous_prob) and (edge_probability < remove_prob):

                #Describing the edge to be removed
                edge_to_remove = [edge[0], edge[1], G[edge[0]][edge[1]]['typ']]

                #Assigning the edge to be removed a unique ID
                real_edges_to_add_back.append([extension_ID, edge_to_remove])

                #Selecting a non-existent edge
                fake_edge = random.choice(edges_not_in_G)
                
                #Picking a random number to assign positive/negative, similar to network creation
                chance = random.random()
                if chance > positive_probability: regulation = '-'
                else: regulation = '+'

                #The new fake edge
                fake_edges_to_add_back.append([fake_edge_ID, [fake_edge[0], fake_edge[1], regulation]])

                #Actually remove the edge from G
                G.remove_edge(edge[0], edge[1])

                #Increment extension IDs
                extension_ID += 1
                fake_edge_ID += 1

        # output_folder = 'ExampleOutput/2_Sample_Removal'
        # model_name = 'network_full.xlsx'

        # #Save the network, the extensions, and a summary
        # #Getting the core model name
        # core_model_name =  model_name.split('.')[0]

        #Percent is just nicer to read
        percent = int(remove_prob*100)

        #Creating Key File Names
        base = output_folder + '/Missing_' + str(percent) + '_percent/' + model_name +'_missing_' + str(percent) + '_percent'
        new_model_name = base + '.xlsx'
        summary_name = base + '.txt'
        real_extensions_name = base + '_grouped_real_ext'
        fake_extensions_name = base + '_grouped_fake_ext'
        figure_name = base + '.png'

        #Make sure directory is created
        os.makedirs(output_folder, exist_ok=True) 
        os.makedirs(output_folder + '/Missing_' + str(percent) + '_percent/', exist_ok=True)

        #Creating the new model file
        network_to_excel(G, filename = new_model_name)
        
        #Make sure all nodes are in the file, even if they have no connections
        extension_network_to_excel(new_model_name, expected_end_values)

        #Creating summary file
        return_extension_stats(base, G, real_edges_to_add_back, fake_edges_to_add_back, verbose = False, save = True, filename = summary_name)

        #Saving the real extensions
        with open(real_extensions_name+'.txt','w') as filer: filer.write(str(real_edges_to_add_back))
        pickle.dump(real_edges_to_add_back,open(real_extensions_name,"wb"))

        #Saving the fake extensions
        with open(fake_extensions_name+'.txt','w') as filer: filer.write(str(fake_edges_to_add_back))
        pickle.dump(fake_edges_to_add_back,open(fake_extensions_name,"wb"))

        #Creating a figure of the new model
        plot_network(G, save = True, filename = figure_name)        

        #Update the bounds of the next probability range
        previous_prob = remove_prob


def score_actual_against_expected_values(expected_dict, actual_dict):
    """Function to quickly retrieve the expected values for a model's simulation. 
    Used to be "get_diff_score"

    Parameters
    ----------
    expected_dict : dictionary
        A dictionary with keys for each node in the network file, and values for the node's end value after simulation.
        Typically this is the "end_vals" dictionary taken from simulating the original network model. 

    actual_dict : dictionary
        A dictionary with keys for each node in the network file, and values for the node's end value after simulation.
        Typically this is the "end_vals" dictionary taken from simulating a network model with missing edges. 

    Returns
    -------
    score : float
        The absolute error/difference between the expected simulation end values and the actual simulation end values.
    """

    individual_differences = {key: ((expected_dict[key] - actual_dict.get(key, 0))**2)**0.5 for key in expected_dict.keys()}
    score = sum(individual_differences.values())
    return score


def BFA(start_model, possible_extensions, simulation_runs, simulation_length, output_directory):
    """The Breadth First Addition (BFA) extension methodology takes an model with missing information, possible extensions, and a simulation scheme \
        to find the optimal combination of extensions that lowers the score. 

    Parameters
    ----------
    start_model : string
        A name of the network file you want to extend using BFA.  

    possible_extensions : list of lists
        A list of lists in which each list entry contains possible extensions used to extend the model. 

    simulation_runs : integer
        The number of simulations to run. This is necessary as the simulator is stochastic, and you can observe different behavior with each simulation. 
        Example = 10 

    simulation_length : integer
        The number of time-steps or simulation updates performed during a simulation run.
        Example = 100

    output_directory : string 
        The name of the folder in which the extension traces will be saved.     
    """
    #Begin the extension process
    iteration = 0
    start_time = dt.now()

    #Define where key values are stored in excel
    extension_tracker = output_directory + 'ExtensionProgress.xlsx'

    #Make sure key directories exist
    extension_folder = output_directory + 'extension_iteration_' + str(iteration) + '/'
    os.makedirs(extension_folder, exist_ok=True) 

    # Get the correct end values of a correct/perfect/expected/golden model
    correct_end_values = get_model_expected_values(start_model)

    #Redefine the start model as the current best model and move it to the output directory
    current_best_model = output_directory + "final.xlsx"
    os.system("cp "+start_model+" "+current_best_model)

    #Simulate the start_model in order to calculate the error between actual end values and expected end values
    simulate_network(start_model, simulation_runs, simulation_length, output_file = extension_folder+'/start_model_trace.txt')

    #Grab the simulation end values from the trace file
    simulation_end_values = get_simulation_end_values(extension_folder+'/start_model_trace.txt', simulation_runs)

    #Calculate the difference between actual simulation and expectation 
    previous_score = float('inf') 
    current_score = get_score_between_expected_and_actual(correct_end_values, simulation_end_values)

    #Report the start of the extension process
    print("Initial difference between the correct and current model: " + str(current_score))

    #Initialize an extension excel to track progress
    wb = Workbook()
    ws = wb.active
    #Column Headers
    ws.cell(row = 1, column = 2, value = "Iteration")
    ws.cell(row = 1, column = 3, value = "Score")
    ws.cell(row = 1, column = 4, value = "Time")
    #First Entry for simulation of base model (which is to be extended)
    ws.cell(row = 2, column = 2, value = "Initial")
    ws.cell(row = 2, column = 3, value = str(current_score))
    ws.cell(row = 2, column = 4, value = int(0))
    #Saving the file
    wb.save(filename = extension_tracker)

    #Initialize key variables to track extension progress
    already_added_extensions = set() # was called ignore
    diff_trend = [current_score]

    #Iterative addition of extensions until no more extensions reduce the score, or you run out of extensions to add
    while current_score < previous_score and len(possible_extensions) > len(already_added_extensions):
        #Update previous difference such that when we no longer see immediate improvement, we stop
        previous_score = current_score

        #Identification of possible extensions not yet added
        extensions_left = (ext for ext in possible_extensions if ext[0] not in already_added_extensions)

        #Initialize the dictionary to hold scores for each extension
        extension_scores_dict = dict()

        #This is a Multiprocessing piece to add and check extensions efficiently -> currently working out the kinks
        # Parallel(n_jobs=numCores,verbose=50)(delayed(eu.extend_unit_breadth_sing)\
        # (curr_model,ext,ignore,simRun,simLen,outPath) \
        # for ext in extensions_left)

        #Iterate through extensions that have not already been added
        for extension_being_added in extensions_left:
            # print("Current extension: "+str(extension_being_added))
            #Initiate the key filenames for this model+extension combination
            ext_trace = extension_folder + "current_best_with_extension_" + str(extension_being_added[0]) + ".txt"
            ext_model = extension_folder + "current_best_with_extension_" + str(extension_being_added[0]) + ".xlsx"

            #extend the current best model with the extension
            extend_model_file(current_best_model,extension_being_added,ext_model)

            #Simulate the new model+extension combination model
            simulate_network(ext_model, simulation_runs, simulation_length, ext_trace)

            #Get the simulation end values
            extension_end_values = get_simulation_end_values(ext_trace, simulation_runs)

            #Compare against expected values, score the extension, and add score to the score dictionary
            this_extension_score = score_actual_against_expected_values(simulation_end_values,extension_end_values)
            extension_scores_dict[extension_being_added[0]] = this_extension_score
            print("Extension: "+str(extension_being_added)+" has a score of: "+str(this_extension_score))      

        #Find the extension that resulted in the minimum score
        minimum_extension_key = min(extension_scores_dict, key=extension_scores_dict.get) 
        print("Minimum found of: "+ str(extension_scores_dict[minimum_extension_key]) + " with extension " + str(minimum_extension_key))

        #Check to see if the minimum extension score is actually an improvement than the current score
        if extension_scores_dict[minimum_extension_key] < previous_score:
            print("Minimum was less than previous score, saving best extension.")

            current_score = extension_scores_dict[minimum_extension_key]
            diff_trend += [current_score]
            already_added_extensions.add(minimum_extension_key)
            ext_model = extension_folder + "current_best_with_extension_" + str(extension_being_added[0]) + ".xlsx"

            iteration += 1
            save_iteration_model = output_directory + "Iteration_" + str(iteration) + ".xlsx"
            os.system("cp "+ext_model+" "+current_best_model)
            os.system("cp "+ext_model+" "+save_iteration_model)

            #Capture the current time to track 
            gen_time = dt.now()
            total_time = gen_time - start_time

            #Updating the logfile for record purposes
            wb = load_workbook(extension_tracker)
            ws = wb.active
            ws.cell(row = iteration + 2, column = 2, value = str(minimum_extension_key))
            ws.cell(row = iteration + 2, column = 3, value = str(current_score))
            ws.cell(row = iteration + 2, column = 4, value = total_time.total_seconds())
            wb.save(filename = extension_tracker)

        else:
            print("Minimum was NOT less than previous score, saving best extension.")

    #Once the model can no longer improve, save the progression
    print("Diff Trend: ",diff_trend)

    #Record the total time for extension process
    end_time = dt.now()
    print("Start time: " + str(start_time))
    print("  End time: " + str(end_time))
    total_time = end_time - start_time
    print("Total time: " + str(total_time))


def recursive_addition_for_DFA(current_best_model, simulation_end_values, \
                            possible_extensions, already_added_extensions, \
                            simulation_runs, simulation_length, \
                            iteration, current_score, output_directory,\
                            diff_trend, start_time):
    """This recursive function continues to extend a baseline model with the first extenion to improve the model \
        until no extension improves the model. This function is exclusively used by the DFA() function. 

    Parameters
    ----------
    current_best_model : string
        The name of the current network file you want to extend using DFA.

    simulation_end_values : dict
        A dictionary

    possible_extensions : list of lists
        A list of lists in which each list entry contains possible extensions used to extend the model. 

    already_added_extensions : set
        A set of numbers indicating which extensions have already been added to the model.

    simulation_runs : integer
        The number of simulations to run. This is necessary as the simulator is stochastic, and you can observe different behavior with each simulation. 
        Example = 10 

    simulation_length : integer
        The number of time-steps or simulation updates performed during a simulation run.
        Example = 100

    iteration : integer
        A number representing the number of recursions that have gone been gone through. Mostly used in notation/nomenclature.

    current_score : float
        A float indicating the current best score obtained during the extension process. Used to measure if a new extension improves the model.

    output_directory : string 
        The name of the folder in which the extension traces will be saved. 

    diff_trend : list
        A list collecting the incremental improvement in the score. Used to report the steps taken to improve the model.

    start_time : datetime.datetime.now() object
        The time the extension method started. Used to record the total elapsed time.     
    """
    #Define where key values are stored in excel
    extension_tracker = output_directory + 'ExtensionProgress.xlsx'

    #Make sure key directories exist
    extension_folder = output_directory + 'extension_iteration_' + str(iteration+1) + '/'
    os.makedirs(extension_folder, exist_ok=True) 

    #Identification of possible extensions not yet added
    extensions_left = (ext for ext in possible_extensions if ext[0] not in already_added_extensions)

    #Iteratively try extensions that have not yet been tried.
    for extension in extensions_left:
        print("Starting extending the current best model with extension: "+str(extension[0]))

        #Take current best model
        current_best_model = output_directory + "final.xlsx"

        #Initiate the key filenames for this model+extension combination
        ext_trace = extension_folder + "current_best_with_extension_" + str(extension[0]) + ".txt"
        ext_model = extension_folder + "current_best_with_extension_" + str(extension[0]) + ".xlsx"

        #extend the current best model with the extension
        extend_model_file(current_best_model,extension,ext_model)

        #Simulate the new model+extension combination model
        simulate_network(ext_model, simulation_runs, simulation_length, ext_trace)

        #Get the simulation end values
        extension_end_values = get_simulation_end_values(ext_trace, simulation_runs)

        #Compare against expected values, score the extension, and add score to the score dictionary
        new_score = score_actual_against_expected_values(simulation_end_values,extension_end_values)
        print("Extension: "+str(extension)+" has a score of: "+str(new_score))      

        #Is the score better? Make it the new best model and call recursion
        if new_score < current_score:
            print("Extension: "+str(extension)+" is an improvement.")
            
            #Update the current score to reflect the new best score
            current_score = new_score
            diff_trend += [current_score]

            #Add extension to list of already_added_extensions to avoid repetition
            already_added_extensions.add(extension[0])

            #Save the improved model as the current_best_model
            iteration += 1
            save_iteration_model = output_directory + "Iteration_" + str(iteration) + ".xlsx"
            os.system("cp "+ext_model+" "+current_best_model)
            os.system("cp "+ext_model+" "+save_iteration_model)

            #Document the improvement
            #Capture the current time to track 
            gen_time = dt.now()
            total_time = gen_time - start_time

            #Updating the logfile for record purposes
            wb = load_workbook(extension_tracker)
            ws = wb.active
            ws.cell(row = iteration + 2, column = 1, value = str(iteration))
            ws.cell(row = iteration + 2, column = 2, value = str(extension[0]))
            ws.cell(row = iteration + 2, column = 3, value = str(current_score))
            ws.cell(row = iteration + 2, column = 4, value = total_time.total_seconds())
            wb.save(filename = extension_tracker)

            #Call recursion
            recursive_addition_for_DFA(current_best_model, simulation_end_values, \
                            possible_extensions, already_added_extensions, \
                            simulation_runs, simulation_length, \
                            iteration, current_score, output_directory,\
                            diff_trend, start_time)

            #End recursion once no more improvement is possible 
            print("No more extensions have been found that improve the model.")

            #Once the model can no longer improve, save the progression
            print("Diff Trend: ",diff_trend)

            #Record the total time for extension process
            end_time = dt.now()
            total_time = end_time - start_time
            print("Start time: " + str(start_time))
            print("  End time: " + str(end_time))
            print("Total time: " + str(total_time))

            break
        
        #No? Better try the next extension then...
        else:
            continue


def DFA(start_model, possible_extensions, simulation_runs, simulation_length, output_directory):
    """The Depth First Addition (DFA) extension methodology takes a model with missing information, possible extensions, and a simulation scheme \
        to find the optimal combination of extensions that lowers the score. 

    Parameters
    ----------
    start_model : string
        A name of the network file you want to extend using BFA.  

    possible_extensions : list of lists
        A list of lists in which each list entry contains possible extensions used to extend the model. 

    simulation_runs : integer
        The number of simulations to run. This is necessary as the simulator is stochastic, and you can observe different behavior with each simulation. 
        Example = 10 

    simulation_length : integer
        The number of time-steps or simulation updates performed during a simulation run.
        Example = 100

    output_directory : string 
        The name of the folder in which the extension traces will be saved.     
    """
    #Begin the extension process
    iteration = 0
    start_time = dt.now()

    #Define where key values are stored in excel
    extension_tracker = output_directory + 'ExtensionProgress.xlsx'

    #Make sure key directories exist
    extension_folder = output_directory + 'extension_iteration_' + str(iteration) + '/'
    os.makedirs(extension_folder, exist_ok=True) 

    # Get the correct end values of a correct/perfect/expected/golden model
    correct_end_values = get_model_expected_values(start_model)

    #Redefine the start model as the current best model and move it to the output directory
    current_best_model = output_directory + "final.xlsx"
    os.system("cp "+start_model+" "+current_best_model)

    #Simulate the start_model in order to calculate the error between actual end values and expected end values
    simulate_network(start_model, simulation_runs, simulation_length, output_file = extension_folder+'/start_model_trace.txt')

    #Grab the simulation end values from the trace file
    simulation_end_values = get_simulation_end_values(extension_folder+'/start_model_trace.txt', simulation_runs)

    #Calculate the difference between actual simulation and expectation 
    previous_score = float('inf') 
    current_score = get_score_between_expected_and_actual(correct_end_values, simulation_end_values)

    #Report the start of the extension process
    print("Initial difference between the correct and current model: " + str(current_score))

    #Initialize an extension excel to track progress
    wb = Workbook()
    ws = wb.active
    #Column Headers
    ws.cell(row = 1, column = 1, value = "Iteration")
    ws.cell(row = 1, column = 2, value = "Extension Group")
    ws.cell(row = 1, column = 3, value = "Score")
    ws.cell(row = 1, column = 4, value = "Time")
    #First Entry for simulation of base model (which is to be extended)
    ws.cell(row = 2, column = 1, value = "Initial Model")
    ws.cell(row = 2, column = 2, value = "None")
    ws.cell(row = 2, column = 3, value = str(current_score))
    ws.cell(row = 2, column = 4, value = int(0))
    #Saving the file
    wb.save(filename = extension_tracker)

    #Initialize key variables to track extension progress
    already_added_extensions = set() # was called ignore
    diff_trend = [current_score]

    # #Call Recursion
    recursive_addition_for_DFA(current_best_model, simulation_end_values, \
                                possible_extensions, already_added_extensions, \
                                simulation_runs, simulation_length, \
                                iteration, current_score, output_directory,\
                                diff_trend, start_time)



#FUTURE WORK

def extend_unit_breadth_sing(mdl,edges,ignore,simRun,simLen,outPath):
    #This has been temporarily scrapped for a slower, easier approach. Will implement in the future. 

	# Identical to function: extend_unit_depth_sing
	#How this function used to be called
	#This is a Multiprocessing piece to add and check extensions efficiently -> currently working out the kinks
	# Parallel(n_jobs=numCores,verbose=50)(delayed(eu.extend_unit_breadth_sing)\
	# (curr_model,ext,ignore,simRun,simLen,outPath) \
	# for ext in extensions_left)

	# ignore.add(edges[0]) #Unnecessary

	trace = outPath+"~temp_trace_"+str(edges[0])+".txt"
	ext_model = outPath + "~temp_final_" + str(edges[0]) + ".xlsx"
	extend_model(mdl,edges,ext_model)

	model = sim.Manager(ext_model)
	model.run_simulation('ra',simRun,simLen,trace,0,outMode=3)


#DEFUNCT
def sim_and_get_diff(G,outPath,model_name,simRun,simLen,correct_val_dic):
    '''
    '''
    first_name = outPath + '/' + model_name + '.xlsx'
    initial_val_dic = create_initial_val_dictionary(G)
    first_G_to_excel(first_name,G,initial_val_dic)
    prop_col = sim_model_and_get_end_vals(first_name,simRun,simLen,initial_val_dic)
    diff = get_diff_score(correct_val_dic,prop_col)
    return diff


def do_the_sub_stuff(H,data_file,count,edges,edge_num_1,edge_num_2,model_name,outPath,simRun,simLen,correct_val_dic):
    '''
    '''
    F = copy.deepcopy(H)
    count += 1
    print("We are on number %s" %(count))
    print("We are working on %s and %s" %(edge_num_1,edge_num_2))
    F.remove_edge(edges[edge_num_2][0], edges[edge_num_2][1])
    model_name_new = model_name + '_' + str(edge_num_1) + '_' + str(edge_num_2)
    delta = sim_and_get_diff(F,outPath,model_name_new,simRun,simLen,correct_val_dic)
    build = str(edge_num_1+1)+','+str(edge_num_2+1)+','+str(delta)+'\n'
    with open(data_file, "a") as myfile:
        myfile.write(build)
    return 1


def parallelize_get_diff(ext,trace,golden,simRun):
	global ext_to_diff
	trace = outPath+"~temp_trace_"+str(ext[0])+".txt"
	ext_to_diff[ext[0]] = get_diff(trace,golden,simRun)