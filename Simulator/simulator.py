import re
import random
import openpyxl

class Manager(object):
	""" Define a model object
	"""

	def __init__(self, model_file, initial_col):
		""" Initialize the model object using model information from the input file
			Inputs:
				model_file : an excel spreadsheet containing element names (col 1),
					activators (col 2), inhibitors (col 3), max number of states (col 4),
					and initial values (initial_col)
		"""

		self.__getElement = dict()
		self.__updateList = list()
		self.__initial = dict()

		# Load the input file containing elements and regulators
		# TODO: replace hardcoding of the column numbers
		wb = openpyxl.load_workbook(model_file)
		ws = wb.active

		# Parse each row of the input file
		# each row contains an element, its regulators, max states, and intial values
		# TODO: also delays?
		curr_row = 2
		while ws.cell(row=curr_row, column=1).value != None:

			# Get the max number of states for each element from the input file
			# Default is 2
			max_state = 3
			if ws.cell(row=curr_row, column=4).value != None:
				max_state = ws.cell(row=curr_row, column=4).value

			# Get spontaneous activation/inhibition delays from input file
			# Default is 1 for each
			act_delay = 1
			inh_delay = 1
			if ws.cell(row=curr_row,column=5).value != None:
				delays = [x.strip() for x in str(ws.cell(row=curr_row,column=5).value).split(',')]
				if len(delays) == 2:
					act_delay = int(delays[0])
					inh_delay = int(delays[1])
				else:
					raise ValueError('Delays in input file must be in the format: \
						[Activation Delay],[Inhibition Delay] \n For example: 1,1')

			# Get initial value from input file
			# Default is 1
			val = 1
			if ws.cell(row=curr_row,column=initial_col).value != None:
				val = ws.cell(row=curr_row,column=initial_col).value

			# Get names of the element (X), activators (A), and inhibitors (I)
			# from input file
			X = ws.cell(row=curr_row,column=1).value
			A = ws.cell(row=curr_row,column=2).value
			I = ws.cell(row=curr_row,column=3).value

			X = '' if X==None else X.strip()
			A = '' if A==None else A.strip()
			I = '' if I==None else I.strip()

			# Create a node object for this element
			# and define initial values, regulators for this element
			# in the model object (self)
			ele = gateNode(X,A,I,val,max_state)
			self.__initial[X] = val
			self.__getElement[X] = ele
			if A!='' or I!='':
				self.__updateList += [ele]

			curr_row += 1

	def set_initial(self):
		""" Set the current value of each element (node) in the model
			to its initial value
		"""

		for name in self.__getElement:
			self.__getElement[name].set_value(self.__initial[name])

	def run_simulation(self, simtype, runs, simStep, outName, **kwargs):
		""" Run a simulation!
			Inputs
				simtype : simulation scheme (not currently used)
				TODO: implement more simulation schemes
				runs : number of simulation runs
				simStep : number of simulation steps
				outName : name of output file
		"""

		# Will only write to output file if outMode!=3
		outMode = kwargs["outMode"] if "outMode" in kwargs else 1

		output_file = open(outName,'w')

		# 'freq_sum' will keep a running sum of the value of each element
		# across runs (frequency)
		freq_sum = dict()
		for key in self.__getElement:
			freq_sum[key] = (simStep+1) * [0]
			freq_sum[key][0] = self.__getElement[key].get_value() * runs

		# Perform 'runs' number of simulation runs
		for run in range(runs):
			if outMode!=3:
				output_file.write('Run #'+str(run)+'\n')

			# Set elements to initial values
			self.set_initial()

			# 'memo' will store this element's values for each step
			# in this run (memory)
			memo = dict()
			for key in self.__getElement:
				memo[key] = [self.__getElement[key].get_value()]

			# Perform 'simStep' number of simulation steps
			for step in range(1,simStep+1):
				# Update elements according to the simulation scheme
				# TODO: define functions for more simulation schemes
				self.ra_update()

				for key in self.__getElement:
					# store values for this step
					memo[key] += [self.__getElement[key].get_value()]
					# increment the sum of values across runs
					freq_sum[key][step] += self.__getElement[key].get_value()

			# Write values from this run to the output file
			if outMode!=3:
				for name in sorted(self.__getElement):
					output_file.write(name+' '+' '.join([str(x) for x in memo[name]])+'\n')


		# Write the sum of values across runs (frequency) to the output file
		output_file.write('\nFrequency Summary:\n')
		for name in sorted(self.__getElement):
			# output_file.write(name+' '
			# 	+' '.join([str(x) for x in freq_sum[name]])+'\n')
			# also write max states for each element to output file
			output_file.write(name+'|'+str(self.__getElement[key].get_max_state())+'|'
				+' '+' '.join([str(x) for x in freq_sum[name]])+'\n')


	def run_simulation_checker(self,simtype,simStep,outName):

		output_file = open(outName,'w')
		output_file.write('# time ')
		for key in sorted(self.__getElement):
			output_file.write(key+'_0 '+key+'_1 ')
		output_file.write('step\n')

		self.print_value(output_file,0)
		for step in range(1,simStep+1):
			self.ra_update()
			self.print_value(output_file,step)

		output_file.close()


	def ra_update(self):
		""" Update all elements, using the random asynchronous (ra) scheme
		"""

		update_ele = random.choice(self.__updateList)
		update_ele.update(self.__getElement)

	def print_value(self,output_file,step):

		output_file.write(str(step)+'  ')
		for key in sorted(self.__getElement):
			val = self.__getElement[key].get_value()
			output_file.write(str(val&1)+' ')
			output_file.write(str((val&2)>>1)+' ')
		output_file.write(str(step)+'\n')


####################################################################
class gateNode(object):
	""" Define a node object containing an element, regulators, and states
	"""
	# TODO: create constants for default values
	def __init__(self,X,A,I,curr_val,max_state=3,delay_act=3,delay_inh=3):
		# the regulated (current) element
		self.__regulated = X.strip()
		# activator of the current element
		self.__act = re.sub('\s','',A)
		# inhibitor of the current element
		self.__inh = re.sub('\s','',I)
		# list of the element and its regulators
		self.__name_list = self.create_name_list(X.strip(),A.strip(),I.strip())
		# dictionary mapping element names to their values
		self.__name_to_value = dict()
		# number of discrete levels (states) for each element
		self.__max_state = int(max_state)
		# current element value
		self.__curr_val = curr_val
		# delay values for spontaneous activation and inhibition
		self.__delay_act = delay_act
		self.__delay_inh = delay_inh
		# current delay values for spontaneous activation and inhibition
		self.__curr_delay_act = 0
		self.__curr_delay_inh = 0

	##### Get functions #####

	def get_name(self):
		return self.__regulated

	def get_max_state(self):
		return self.__max_state

	def get_name_list(self):
		return self.__name_list

	def get_value(self):
		return self.__curr_val

	#########################

	def set_value(self,val):
		""" Set this element's current value """
		self.__curr_val = val

	def create_name_list(self,X,A,I):
		""" Create a list of this element, activator, and inhibitor names """
		names = set([X])
		act_set = set(re.findall(r'[\.\w_@\*\#/;]+',A))
		inh_set = set(re.findall(r'[\.\w_@\*\#/;]+',I))
		# TODO: subtracting names?
		return sorted(list(act_set-names)) + sorted(list(inh_set-act_set-names)) + list(names)

	def update(self,getElement):
		""" Update this element """
		self.__name_to_value.clear()
		for name in self.__name_list:
			self.__name_to_value[name] = getElement[name].get_value()
		self.__curr_val = self.evaluate()

	def evaluate(self):
		""" determine the value of the regulated element
			based on values of the inhibitors and activators
		"""

		# calculate activation and inhibition scores
		y_act = self.eval_act(self.__act,0)
		y_inh = self.eval_inh(self.__inh,0)

		# define current element value, max states, and max delays for code readability
		X_curr = int(self.__name_to_value[self.__regulated])
		N = self.__max_state-1
		D_act = self.__delay_act
		D_inh = self.__delay_inh

		# determine next value of the regulated element,
		# based on the type of regulators and activation/inhibition scores
		if (self.__act) and (not self.__inh):
			# this element has only positive regulators
			# set the next value to increase if activation > 0,
			# but don't increase beyond the maximum state.
			# if there is no activation, the value should spontaneously decay,
			# but not below 0
			if (y_act > 0) and (X_curr < N):
				X_next = X_curr + 1
			elif (y_act > 0) and (X_curr == N):
				X_next = X_curr
			elif (y_act == 0) and (X_curr > 0):
				# spontaneously decay (inhibition)
				if self.__curr_delay_inh < D_inh:
					# hold current value and increment delay
					X_next = X_curr
					self.__curr_delay_inh += 1
				else:
					# decay and reset delay
					X_next = X_curr - 1
					self.__curr_delay_inh = 0
			elif (y_act == 0) and (X_curr == 0):
				X_next = X_curr
		elif (not self.__act) and (self.__inh):
			# this element has only negative regulators
			# set the next value to decrease if inhibition > 0,
			# but don't decrease below 0.
			# if there is no inhibition, the value should spontaneously increase,
			# but not beyond the maximum state.
			if (y_inh > 0) and (X_curr > 0):
				X_next = X_curr - 1
			elif (y_inh > 0) and (X_curr == 0):
				X_next = X_curr
			elif (y_inh == 0) and (X_curr < N):
				# spontaneously increase
				if self.__curr_delay_act < D_act:
					# hold current value and increment delay
					X_next = X_curr
					self.__curr_delay_act += 1
				else:
					# increase and reset delay
					X_next = X_curr + 1
					self.__curr_delay_act = 0
			elif (y_inh == 0) and (X_curr == N):
				X_next = X_curr
		elif (self.__act) and (self.__inh):
			# this element has both activators and inhibitors
			# increase the value if activation > inhibition,
			# decrease if activation <= inhibition,
			# but not above/below the bounds of the max/min states.
			# Note: this means if activation == inhibition,
			# the value spontaneously decays.
			if 	(y_act > y_inh) and (X_curr < N):
				X_next = X_curr + 1
			elif (y_act > y_inh) and (X_curr == N):
				X_next = X_curr
			elif (y_act <= y_inh) and (X_curr > 0):
				# spontaneously decay (inhibition)
				if self.__curr_delay_inh < D_inh:
					# hold current value and increment delay
					X_next = X_curr
					self.__curr_delay_inh += 1
				else:
					# decay and reset delay
					X_next = X_curr - 1
					self.__curr_delay_inh = 0
			elif (y_act <= y_inh) and (X_curr == 0):
				X_next = X_curr

		elif (not self.__act) and (not self.__inh):
			# this element has no regulators; hold its value
			X_next = X_curr

		# return the next state,
		# with a redundant sort to keep X_next within the state value bounds
		return sorted([0, X_next , N])[1]

	def eval_inh(self,inh_rule,layer):
		""" returns an int y_sum indicating the inhibition score
			calculated based on the value of the inhibitors
			Inputs:
				inh_rule : inhibitor function notation
				layer : set != 0 when the function is called recursively
		"""

		# Only calculate the score if there are actually inhibitors for this element
		if inh_rule:
			# create a list of inhibitors from influence set notation
			# inhibitors are separated by commas (outside parentheses)
			inh_list = self.split_comma_outside_parentheses(inh_rule)
			y_sum = list()

			# parse inhibitors, first checking for groups of elements in ()
			for inh_element in inh_list:
				if inh_element[0]=='(' and inh_element[-1]==')':
					# this is and AND operation, all inhibitors must be present
					# construct a list of the values of each element, then perform discrete AND (min)
					y_and = [int(x) \
						for and_entity in self.split_comma_outside_parentheses(inh_element[1:-1]) \
						for x in self.eval_inh(and_entity,1)]
					y_sum += [min(y_and)]
				else:
					# single inhibitor
					# confirm there are no commas separating inhibitors
					assert(inh_element.find(',')==-1)
					# TODO: add code below for highest state inhibitors?
					if inh_element[-1]=='+':
					# highest state inhibitor
					# increment the sum, calculating the increment value by comparing to the max state
						if inh_element[0]=='!':
							y_sum += int(not self.__name_to_value[inh_element[1:-1]]==self.__max_state-1)
						else:
							y_sum += int(self.__name_to_value[inh_element[:-1]]==self.__max_state-1)*2
					elif inh_element[0]=='!':
						y_sum += [int(self.discrete_not(self.__name_to_value[inh_element[1:]],self.__max_state-1))]
					else:

					# calculate the value of the sum based on the value of the inhibitor
					# using the state name to value mapping (__name_to_value dictionary)
						y_sum += [int(self.__name_to_value[inh_element])]

			if layer==0:
				# TODO: initializer values for inhibitors?
				return max(y_sum)
			else:
				return y_sum

	def eval_act(self,act_rule,layer):
		""" returns a string list y_sum indicating the activation score,
			calculated based on the value of the activators
			Inputs:
				act_rule : activator function notation
				layer : set != 0 when the function is called recursively
		"""

		# Only calculate the score if there are actually activators for this element
		if act_rule:
			# TODO: define N = self.__max_state-1 to make more readable

			# create a list of activators from influence set notation
			# activators are separated by commas (outside parentheses)
			act_list = self.split_comma_outside_parentheses(act_rule)

			y_init = list()
			y_sufficient = list()
			# TODO: explain why score is initialized as a list
			y_sum = list()
			y_must = list()
			y_enhance = list()
			# parse activators, first checking for elements in {},  ()
			for act_element in act_list:
				if act_element[0]=='{' and act_element[-1]=='}':
					# this is an initializer
					# confirm that this is layer 0
					assert(layer==0)
					# define as an initializer, evaluating only the expression within the brackets
					y_sum += self.eval_act(act_element[1:-1],1)
				elif act_element[0]=='{' and act_element[-1]==']':
					# this is a necessary pair
					# Find the cut point between {} and []
					parentheses = 0
					cut_point = 0
					for index in range(len(act_element)):
						if act_element[index]=='{':
							parentheses += 1
						elif act_element[index]=='}':
							parentheses -= 1
						if parentheses==0:
							cut_point = index
							break
					# define the first part as the sufficient element
					y_must += self.eval_act(act_element[1:cut_point],1)
					# define the second part as the enhancing/strengthening element
					y_enhance += self.eval_act(act_element[cut_point+2:-1],1)
					# increment the score according to the values of both the sufficient and enhancing elements
					# but use the 'sorted' expression to keep the value below the max state value
					# TODO: check min(min(y_must),max(y_enhance))
					y_sum += [0 if all([y==0 for y in y_must])==True \
						else sorted([0, int(max(min(y_must),max(y_enhance))), self.__max_state-1])[1]]
				elif act_element[0]=='(' and act_element[-1]==')':
					# this is an AND operation, all activators must be present
					# construct a list of the values of each element, then perform discrete AND (min)
					y_and = [int(x) \
						for and_entity in self.split_comma_outside_parentheses(act_element[1:-1]) \
						for x in self.eval_act(and_entity,1)]
					y_sum += [min(y_and)]
				else:
					# single activator
					# confirm that there are no commas separating activators
					assert(act_element.find(',')==-1)
					# calculate the value of the score based on the value of the activator
					# using the state name to value mapping (__name_to_value dictionary)
					if act_element[-1]=='+':
						# this is a highest state activator
						# increment the score only if this activator is at the max state
						if act_element[0]=='!':
							y_sum += [int(not self.__name_to_value[act_element[1:-1]]==self.__max_state-1)]
						else:
							# TODO: Why multiplying by 2? check Description.pptx for scoring info
							y_sum += [int(self.__name_to_value[act_element[:-1]]==self.__max_state-1)*2]
					elif act_element[0]=='!':
						y_sum += [int(self.discrete_not(self.__name_to_value[act_element[1:]],self.__max_state-1))]
					else:
						y_sum += [int(self.__name_to_value[act_element])]

			if layer==0:
				# check for initializers and value of intializers
				if (self.__name_to_value[self.__regulated]==0
					and len(y_init)!=0
					and all([y==0 for y in y_init])==True):
					return 0
				else:
					return max(y_sum)

			else:
				return y_sum

	def discrete_not(self, x, N):
		"""
		"""
		if N >= x:
			return (N - x)
		else:
			raise ValueError('Can''t compute NOT, input is greater than max state')

	def split_comma_outside_parentheses(self, sentence):
		final_list = list()
		parentheses = 0
		start = 0
		for index in range(len(sentence)):
			char = sentence[index]
			if index==len(sentence)-1:
				final_list.append(sentence[start:index+1])
			elif char=='(' or char=='{' or char=='[':
				parentheses += 1
			elif char==')' or char=='}' or char==']':
				parentheses -= 1
			elif (char==',' and parentheses==0):
				final_list.append(sentence[start:index])
				start = index+1
		return final_list

	# TODO: add the function to generate a model expression (same as in utility_function.py)

# TODO: define a main function?
# def main():

# if __name__=="__main__":
# 	main()
