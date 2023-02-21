import re
import random
import openpyxl
import math

####################################################################
############ 			Manager object  		 		############
####################################################################
class Manager(object):
	""" Define a simulation manager object, containing simulation parameters and elements. 

		The Manager object contains dictionaries of: 
			- model element objects (gateNode objects)
			- initial values
			- toggle times and values
			- update groups
			- TODO: update probabilities
			- TODO: update ranks 
		Any element-specific model characteristics are properties of the element objects (gateNode)
	"""

	def __init__(self, model_file,initial_col):
		""" Initialize the model object using model information from the input file
			Inputs:
				model_file : an excel spreadsheet containing element names (col 1),
					activators (col 2), inhibitors (col 3), max number of states (col 4),
					and initial values (col 6)
		"""

		# Model Defaults (TODO: make inputs to the __init__ function? check issues with mutable types)
		DEF_MAX_STATE = 3
		DEF_SPONT_ACT = 2
		DEF_SPONT_INH = 2

		# Input file columns (TODO: use column names to index from dictionary file)
		input_col_X = 1
		input_col_A = 2
		input_col_I = 3
		input_col_maxstate = 4
		input_col_delays = 5
		input_col_initial = initial_col
		input_col_spont = 7
		input_col_toggle_time= 8
		input_col_toggle_val = 9
		input_col_update_group = 10

		# Initialize model properties
		# List of all elements in the model
		self.__getElement = dict()
		# List of elements with positive and/or negative regulators
		self.__updateList = list()
		# Initial values for each element
		self.__initial = dict()
		# Max state across all elements (TODO: remove this, just use max state for each element)
		self.__maxstate = DEF_MAX_STATE
		# List of elements that toggle
		self.__switchList = dict()
		# Toggle values for toggle elements
		self.__switchValue = dict()
		# Element group values for group updating
		self.__groupUpdate = dict()

		# Load the input file containing elements and regulators
		wb = openpyxl.load_workbook(model_file)
		ws = wb.active		

		# Parse each row of the input file 
		# each row contains an element, its regulators, max states, delays,
		# initial values, and update group
		curr_row = 2
		while ws.cell(row=curr_row, column=input_col_X).value is not None:
			# Each row in the spreadsheet corresponds to one element

			# Get names of the element (X), activators (A), and inhibitors (I) 
			X = ws.cell(row=curr_row,column=input_col_X).value
			A = ws.cell(row=curr_row,column=input_col_A).value
			I = ws.cell(row=curr_row,column=input_col_I).value
			
			X = '' if X==None else X.strip()
			A = '' if A==None else A.strip()
			I = '' if I==None else I.strip()

			# Get the max number of states for this element 
			# if different from the default max state for the model
			# TODO: set the max model state according to the max state across elements instead of having a model default
			max_state = self.__maxstate
			if ws.cell(row=curr_row, column=input_col_maxstate).value is not None:
				max_state = ws.cell(row=curr_row, column=input_col_maxstate).value

			# Define element level transition delays; 
			# different delays can be defined for each level transition (rising or falling edge)
			delays = [0 for x in range(2*(max_state-1))]
			if ws.cell(row=curr_row,column=input_col_delays).value is not None:
				delays = [int(x.strip()) for x in str(ws.cell(row=curr_row,column=input_col_delays).value).split(',')]
			# check that the number of delays matches the number of states
			if len(delays) != 2*(max_state-1):
				if len(delays) == 1:
					# use the same delay for all transitions
					delays = [int(delays[0]) for x in range(2*(max_state-1))]
				else:
					# incorrect delay format in the input file
					raise ValueError('For element delays, need 2*(max_state - 1) delays in the format: \n \
						delay01,delay12,delay21,delay10 \n \
						For example, for a 3 state model (levels 0, 1, 2): 1,2,2,1 \n \
						OR, only one delay value, which will be used for all level transitions \n \
						OR, leave blank for no delay')

			# Get initial value for this element
			initial_value_input = ws.cell(row=curr_row, column=input_col_initial).value
			if str(initial_value_input).lower().strip() in ['r','random']:
				# TODO: move this randomization to just before a simulation is run, so multiple
				# simulations can be run with random initialization without having to re-create the model
				# Here, just set self.__initial[ele] = 'r' or something to indicate it should be random 
				r = random.randrange(3)
				if r == 0:
					val = 0
				elif r == 1:
					val = int(max_state/2)
				elif r == 2:
					val = (max_state-1)  
			elif str(initial_value_input).lower().strip() in ['l','low']:
				val = 0
			elif str(initial_value_input).lower().strip() in ['m','med','medium','middle','moderate']:
				val = int(max_state/2)
			elif str(initial_value_input).lower().strip() in ['h','high']:
				val = (max_state-1) 
			elif initial_value_input is not None:
				val = initial_value_input
			else:
				raise ValueError('Missing initial value for element ' + str(X))

			# Define spontaneous activation/inhibition behavior and delays 
			# TODO: use spontaneously restores (R) or degrades (D) from dictionary
			spont_act = DEF_SPONT_ACT
			spont_inh = DEF_SPONT_INH
			if ws.cell(row=curr_row,column=input_col_spont).value is not None:
				spont_delays = [x.strip() for x in str(ws.cell(row=curr_row,column=input_col_spont).value).split(',')]
				if len(spont_delays) == 2:
					spont_act = int(spont_delays[0]) if spont_delays[0] != '' else ''
					spont_inh = int(spont_delays[1]) if spont_delays[1] != '' else ''
				else:
					raise ValueError('Spontaneous behavior in the input file must be in the format: \
						[Activation Delay],[Inhibition Delay] \n \
						For example: 1,1 \n \
						If spontaneous behavior is desired with no delay, input 0,0 \n \
						If only spontaneous inhibition is desired, input ,0 \n \
						If no spontaneous behavior is desired, input ,')

			# Create a node object for this element
			# set the intial values, define regulators and behavior characteristics for this element
			ele = gateNode(X, A, I, val, max_state, delays, spont_act, spont_inh)

			# Include this element in the model's dictionary of elements
			self.__getElement[X] = ele

			# Set model or simulation-related values for this element
			# e.g., initial value, toggle time and value, update group, etc.
			self.__initial[X] = val

			# Set toggle time and value for this element
			if ws.cell(row=curr_row, column=input_col_toggle_time).value is not None:
				S = str(ws.cell(row=curr_row, column=input_col_toggle_time).value)
				S = S.split(".")
				S = ' '.join(S).split()
				for i in range(len(S)):
					S[i] = int(S[i])
				self.__switchList[X] = S
			if ws.cell(row=curr_row, column=input_col_toggle_val).value is not None:
			    V = str(ws.cell(row=curr_row, column=input_col_toggle_val).value)
			    V = V.split(".")
			    V = ' '.join(V).split()
			    for i in range(len(V)):
			        V[i] = int(V[i])
			    self.__switchValue[X] = V

			# Set this element's update group for simulations
			if ws.cell(row=curr_row, column=input_col_update_group).value is not None:
			    group = ws.cell(row=curr_row, column=input_col_update_group).value
			    self.__groupUpdate[X] = group

			# Add this element to the list of elements that will be updated if it has regulators
			if A!='' or I!='':
				self.__updateList += [ele]

			# Go to the next row in the input file
			curr_row += 1

	## Get/set functions
	
	def get_elements(self):
		return self.__getElement

	def get_initial(self):
		return self.__initial

	def set_initial(self): #set_initial(self,scheme):
		""" Set the current value of each element (node) in the model
			to its initial value
		"""
		for name in self.__getElement:
			# TODO: add code below (and add scheme function input) to randomize initial value 
			# if self.__initial[name] == 'r' or scheme == 'sync':
				# r = random.randrange(3)
				# if r == 0:
				#     val = 0
				# if r == 1:
				#     val = int(self.__getElement[name].get_max_state()/2)
				# if r == 2:
				#     val = self.__getElement[name].get_max_state()-1
				# self.__getElement[name].set_value(val)
			# else:
			val = self.__initial[name]
			self.__getElement[name].set_value(val)

	def set_initial_sync(self):
		""" Randomize the initial values of elements for the synchronous/simultaneous
			scheme to low, medium, or high based on the max state
		"""
		# Even though the initial state is randomized when the Manager is initialized,
		# this function is needed so you can run multiple synchronous simulations
		# with random initial values using the same Manager object
		# TODO: use this functionality in all simulations (in set_initial) to allow random initializations of some values
		for name in self.__getElement:
		    r = random.randrange(3)
		    if r == 0:
		        valeur = 0
		    if r == 1:
		        valeur = int(self.__maxstate/2)
		    if r == 2:
		        valeur = self.__maxstate-1
		    self.__getElement[name].set_value(valeur)

	## Simulation functions

	def run_simulation(self, simtype, runs, simStep, outName, **kwargs):
		""" Run a simulation!
			Inputs
				simtype : simulation scheme 
					'ra' : random asynchronous
					'sync' : synchronous / simultaneous
					TODO: update to new terminology
				runs : number of simulation runs
				simStep : number of simulation steps
				outName : name of output file
				outMode : specify output mode (1: Original Version, 2: Model Checker, 3: Summary Only)
		"""

		# Will write to output file according to outMode
		outMode = kwargs["outMode"] if "outMode" in kwargs else 1

		output_file = open(outName,'w')
		
		# 'freq_sum' will keep a running sum of the value of each element 
		# across runs (frequency)
		freq_sum = dict()
		for key in self.__getElement:
			freq_sum[key] = (simStep+1) * [0]
			freq_sum[key][0] = self.__getElement[key].get_value() * runs
		
		# Perform 'runs' number of simulation runs
		for run in range(runs):
			if outMode==1: 
				output_file.write('Run #'+str(run)+'\n')
			
			# Set elements to initial values
			self.set_initial()

			# 'memo' will store this element's values for each step 
			# in this run (memory)
			memo = dict()
			for key in self.__getElement:
				memo[key] = [self.__getElement[key].get_value()]

			# Perform 'simStep' number of simulation steps
			for step in range(1,simStep+1):
				# Update elements according to the simulation scheme
				if simtype == 'ra' :
					# randomly choose an element or element group to update
					name = self.ra_update()
					if name in self.__groupUpdate:
					    for key in self.__groupUpdate:
					        if self.__groupUpdate[name] == self.__groupUpdate[key]:
					            self.update(key)
				if simtype == 'sync':
					# simultaneously update all elements
					self.sync_update()

				for key in self.__getElement:
					# store values for this step
					memo[key] += [self.__getElement[key].get_value()]
					# increment the sum of values across runs
					freq_sum[key][step] += self.__getElement[key].get_value()

					# Check for element value toggles
					if key in self.__switchList:
						for i in range(len(self.__switchList[key])):
							if self.__switchList[key][i] == step:
								memo[key][step] = self.__switchValue[key][i]
								freq_sum[key][step] = self.__switchValue[key][i] * runs
								self.__getElement[key].set_value(self.__switchValue[key][i])

					# TODO: increment delays here if they should be updated every time step

			# Write values from this run to the output file
			if outMode==1: 
				for name in sorted(self.__getElement):
					output_file.write(name+' '+' '.join([str(x) for x in memo[name]])+'\n')
			elif outMode==2:
				# transpose format (used by sensitivity analysis and model checking)
				if run == 0: 
					output_file.write('# time ')
					output_file.write(' '.join([name for name in sorted(self.__getElement)]))
					output_file.write(' step\n')
				
				for step in range(simStep):
					output_file.write(str(step)+'  ')
					output_file.write(' '.join([str(memo[name][step]) for name in sorted(self.__getElement)]))
					output_file.write(' '+str(step)+'\n')

		# Write frequency summary (sum of values for each element at each step across runs)	
		if outMode != 2:
			if outMode==3:
				# Write total number of runs, to be used for plotting  
				output_file.write('Run #'+str(run)+'\n')

			output_file.write('\nFrequency Summary:\n')
			for name in sorted(self.__getElement):
				# also write max states for each element to output file so they can 
				# be used to plot the traces later
				output_file.write(name+'|'+str(self.__getElement[name].get_max_state())+'|'
					+' '+' '.join([str(x) for x in freq_sum[name]])+'\n')


	def update(self,element):
		for i in range(len(self.__updateList)):
			if element == self.__updateList[i].get_name():
			    update_ele = self.__updateList[i]
			    update_ele.update(self.__getElement)
		
	def ra_update(self):
		""" Update all elements, using the random asynchronous (ra) scheme
		"""

		update_ele = random.choice(self.__updateList)
		update_ele.update(self.__getElement)
		return update_ele.get_name()

	def sync_update(self):
	    i = 0
	    while i < len(self.__updateList):
	        update_ele = self.__updateList[i]
	        update_ele.update(self.__getElement)
	        i += 1

	def run_simulation_checker(self,simtype,simStep,outName):
		# This writes an output trace file that can be used with 
		# other code that needs boolean values (e.g., sensitivity analysis)
		# TODO: rename this function to indicate that it just outputs a trace (but first check to see if any other code is using it)
		output_file = open(outName,'w')
		output_file.write('# time ')

		# this code is assuming a max of 2 bits to represent an element's levels
		# so it makes _0 and _1 variables for each element,
		# TODO: update to convert any number of discrete levels to boolean variables
		for key in sorted(self.__getElement):
			output_file.write(key+'_0 '+key+'_1 ')
		output_file.write('step\n')

		self.print_value(output_file,0)
		for step in range(1,simStep+1):
			if simtype == 'ra':
			    self.ra_update()
			if simtype == 'sync':
			    self.sync_update()
			self.print_value(output_file,step)

		output_file.close()		

	def print_value(self,output_file,step):
		# Used by run_simulation_checker to convert values to boolean,
		# this code is assuming a max of 2 bits to represent an element's levels
		# TODO: update for any number of discrete levels
		# TODO: rename this function to indicate that it converts values to Boolean (but first check if any other code is using it)
		output_file.write(str(step)+'  ')
		for key in sorted(self.__getElement):
			val = self.__getElement[key].get_value()
			output_file.write(str(val&1)+' ')
			output_file.write(str((val&2)>>1)+' ')
		output_file.write(str(step)+'\n')

	
####################################################################
############ 			gateNode object  		 		############
####################################################################
class gateNode(object):
	""" Define a node object representing an element.
		Each element object has properties that include:
			activators,
			inhibitors,
			number of activity levels,
			delay behavior and TODO: update function (min/max, sum, truth table),
			current value
			
	"""
	
	# Note: no default values are provided here to avoid conflicts with the default
	# values for a model object (Manager)
	def __init__(self, X, A, I, curr_val, max_state, delays, spont_act, spont_inh):
		# the regulated (current) element name
		self.__regulated = X.strip()

		# names of activators of this element
		self.__act = re.sub('\s','',A)
		# names of inhibitors of this element
		self.__inh = re.sub('\s','',I)

		# number of discrete levels (states) for each element
		self.__max_state = int(max_state)

		# current element value
		self.__curr_val = curr_val

		# delay values for spontaneous activation and inhibition
		self.__spont_act = spont_act
		self.__spont_inh = spont_inh
		# current delay values for spontaneous activation and inhibition
		self.__curr_spont_delay_act = 0
		self.__curr_spont_delay_inh = 0
		# delay values for element level transition delays
		self.__delays = delays
		self.__max_delay = max(delays)

		# current delay values for element level transition delays
		self.__curr_delays = [0 for x in delays]

		# names of this element and its regulators
		self.__name_list = self.create_name_list(X.strip(),A.strip(),I.strip())

		# dictionary mapping the names of this element and its regulators to their values
		self.__name_to_value = dict()

	##############################################################
	# Get/set functions
	##############################################################

	def get_name(self):
		return self.__regulated

	def get_max_state(self):
		return self.__max_state

	def get_name_list(self):
		return self.__name_list

	def get_value(self):
		return self.__curr_val

	def get_delays(self):
		return self.__delays

	def set_value(self,val):
		""" Set this element's value """
		self.__curr_val = val

	#############################################################

	def create_name_list(self,X,A,I):
		""" Create a list of this element, activator, and inhibitor names """
		names = set([X])
		act_set = set(re.findall(r'[\.\w_@\*\#/;]+',A))
		inh_set = set(re.findall(r'[\.\w_@\*\#/;]+',I))
		# Note: subtracting names removes repeats of X due to self-regulation
		return sorted(list(act_set-names)) + sorted(list(inh_set-act_set-names)) + list(names)

	def update(self,getElement):
		""" Update this element's dictionary of its regulator's values,
			then update the element's value """
		self.__name_to_value.clear()
		# update the state (values of this element and its regulators)
		for name in self.__name_list:
			self.__name_to_value[name] = getElement[name].get_value()
		
		# TODO: call different evaluate functions?
		self.__curr_val = self.evaluate()

	def evaluate(self):
		""" Determine the value of the regulated element based on values of the inhibitors and activators.
			Uses greater/less than comparison of the regulator scores and increments/decrements the element value by 1.
			Incorporates state-transition delays.
			TODO: rename to indicate which update method this uses, add other update methods
		"""

		# calculate activation and inhibition scores
		y_act = self.eval_reg(self.__act,0)
		y_inh = self.eval_reg(self.__inh,0)

		# define current element value, max states, and max delays for code readability
		X_curr = self.__name_to_value[self.__regulated]
		N = self.__max_state-1
		D_act = self.__spont_act
		D_inh = self.__spont_inh
		D = self.__delays

		# determine next value of the regulated element,
		# based on the type of regulators and activation/inhibition scores
		# TODO: condense by reintroducing the gradient value
		# TODO: reset spontaneous delays if there is activation/inhibition?
		# TODO: use difference between y_act and y_inh for strength of increase/decrease in next-state value
		if (self.__act) and (not self.__inh): 
			# this element has only positive regulators
			# set the next value to increase if activation > 0,
			# but don't increase beyond the maximum state.
			# if there is no activation, the value should spontaneously decay, 
			# but not below 0
			if (y_act > 0) and (X_curr < N):
				# check the delay value
				# there can be different delays for each level transition
				# since this is an increase, index the delays list using the current value of X
				# so if X is currently 0, and transitioning from 0 - 1, we want delays[0]
				# therefore, our index is X_curr
				if self.__curr_delays[int(X_curr)] < D[int(X_curr)]:
					# hold current value and increment delay 
					X_next = X_curr
					self.__curr_delays[int(X_curr)] += 1
				else:
					# increase value and reset delay
					X_next = X_curr + 1
					self.__curr_delays[int(X_curr)] = 0

			elif (y_act > 0) and (X_curr == N):
				X_next = X_curr
			elif (y_act == 0) and (X_curr > 0):
				if D_inh != '':
					# spontaneously decay (inhibition)
					if self.__curr_spont_delay_inh < D_inh:
						# hold current value and increment spontaneous delay
						X_next = X_curr
						self.__curr_spont_delay_inh += 1
					else:
						# decay and reset spontaneous delay
						X_next = X_curr - 1
						self.__curr_spont_delay_inh = 0
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_act == 0) and (X_curr == 0):
				X_next = X_curr
		elif (not self.__act) and (self.__inh):
			# this element has only negative regulators
			# set the next value to decrease if inhibition > 0,
			# but don't decrease below 0.
			# if there is no inhibition, the value should spontaneously increase, 
			# but not beyond the maximum state.
			if (y_inh > 0) and (X_curr > 0):
				# check the delay value
				# there can be different delays for each level transition
				# since this is a decrease, negative index the delays list using the current value of X
				# For example:
				# for max_state=3, delays = [delay01, delay12, delay21, delay10]
				# so if X is currently 1, and transitioning from 1-0, we want delays[-1]
				# therefore, our index is -X_curr
				if self.__curr_delays[-int(X_curr)] < D[-int(X_curr)]:
					# hold current value and increment delay 
					X_next = X_curr
					self.__curr_delays[-int(X_curr)] += 1
				else:
					# decrease value and reset delay
					X_next = X_curr - 1
					self.__curr_delays[-int(X_curr)] = 0

			elif (y_inh > 0) and (X_curr == 0):
				X_next = X_curr
			elif (y_inh == 0) and (X_curr < N):
				if D_act != '':
					# spontaneously increase
					if self.__curr_spont_delay_act < D_act:
						# hold current value and increment spontaneous delay
						X_next = X_curr
						self.__curr_spont_delay_act += 1
					else:
						# increase and reset spontaneous delay
						X_next = X_curr + 1
						self.__curr_spont_delay_act = 0
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_inh == 0) and (X_curr == N):
				X_next = X_curr		
		elif (self.__act) and (self.__inh):
			# this element has both activators and inhibitors
			# increase the value if activation > inhibition,
			# decrease if activation <= inhibition,
			# but not above/below the bounds of the max/min states.
			# Note: if activation == inhibition, 
			# the value spontaneously decays.
			if 	(y_act > y_inh) and (X_curr < N):
				# check the delay value
				# there can be different delays for each level transition
				# since this is an increase, index the delays list using the current value of X
				# so if X is currently 0, and transitioning from 0 - 1, we want delays[0]
				# therefore, our index is X_curr
				if self.__curr_delays[int(X_curr)] < D[int(X_curr)]:
					# hold current value and increment delay 
					X_next = X_curr
					self.__curr_delays[int(X_curr)] += 1
				else:
					# increase value and reset delay
					X_next = X_curr + 1
					self.__curr_delays[int(X_curr)] = 0

			elif (y_act > y_inh) and (X_curr == N):
				X_next = X_curr
			elif (y_act == y_inh) and (X_curr > 0):
				if D_inh != '':
					# spontaneously decay (inhibition)
					if self.__curr_spont_delay_inh < D_inh:
						# hold current value and increment spontaneous delay
						X_next = X_curr
						self.__curr_spont_delay_inh += 1
					else:
						# decay and reset spontaneous delay
						X_next = X_curr - 1
						self.__curr_spont_delay_inh = 0
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_act < y_inh) and (X_curr > 0):
				# check the delay value
				# there can be different delays for each level transition
				# since this is a decrease, negative index the delays list using the current value of X
				# For example:
				# for max_state=3, delays = [delay01, delay12, delay21, delay10]
				# so if X is currently 1, and transitioning from 1-0, we want delays[-1]
				# therefore, our index is -X_curr
				if self.__curr_delays[-int(X_curr)] < D[-int(X_curr)]:
					# hold current value and increment delay 
					X_next = X_curr
					self.__curr_delays[-int(X_curr)] += 1
				else:
					# decrease value and reset delay
					X_next = X_curr - 1
					self.__curr_delays[-int(X_curr)] = 0

			elif (y_act <= y_inh) and (X_curr == 0):
				X_next = X_curr
		else:
			# this element has no regulators; 
			# Note that this shouldn't happen with the current model initialization
			X_next = X_curr
		
		# return the next state,
		# with a redundant sort to keep X_next within the state value bounds
		return sorted([0, X_next , N])[1]

	def eval_reg(self,reg_rule,layer):
		""" Returns a string list y_sum indicating the regulator score,
			calculated based on the value of the activators or inhibitors (the list in reg_rule).
			Uses discrete AND, OR, and NOT (min, max, n's complement).
			Inputs:
				reg_rule : activator or inhibitor function notation 
				layer : set != 0 when the function is called recursively
		"""
		
		# Only calculate the score if there are actually activators for this element
		if reg_rule:

			N = self.__max_state-1 

			# create a list of activators from influence set notation
			# activators are separated by commas (outside parentheses)
			reg_list = self.split_comma_outside_parentheses(reg_rule)

			y_init = list()
			y_sufficient = list() 
			y_sum = list()
			y_must = list()
			y_enhance = list()

			# parse activators, first checking for elements in {},  ()
			for reg_element in reg_list:
				if reg_element[0]=='{' and reg_element[-1]=='}':
					# this is an initializer
					# confirm that this is layer 0
					assert(layer==0)
					# define as an initializer, evaluating only the expression within the brackets
					y_sum += self.eval_reg(reg_element[1:-1],1)
				elif reg_element[0]=='{' and reg_element[-1]==']':
					# this is a necessary pair
					# Find the cut point between {} and []
					parentheses = 0
					cut_point = 0
					for index in range(len(reg_element)): 
						if reg_element[index]=='{':
							parentheses += 1
						elif reg_element[index]=='}':
							parentheses -= 1
						if parentheses==0:
							cut_point = index
							break
					# define the first part as the sufficient element
					y_must += self.eval_reg(reg_element[1:cut_point],1)
					# define the second part as the enhancing/strengthening element
					y_enhance += self.eval_reg(reg_element[cut_point+2:-1],1)
					# increment the score according to the values of both the sufficient and enhancing elements
					# but use the 'sorted' expression to keep the value below the max state value
					# TODO: choose AND/OR functions 
					y_sum += [0 if all([y==0 for y in y_must])==True \
						else sorted([0, max(y_must)+max(y_enhance), N])[1]]
				elif reg_element[0]=='(' and reg_element[-1]==')':
					# this is an AND operation, all activators must be present
					# construct a list of the values of each element, then perform discrete AND (min)
					y_and = [int(x) \
						for and_entity in self.split_comma_outside_parentheses(reg_element[1:-1]) \
						for x in self.eval_reg(and_entity,1)]
					# TODO: choose AND function
					y_sum += [min(y_and)]
				else:
					# single activator
					# confirm that there are no commas separating activators
					assert(reg_element.find(',')==-1)
					# calculate the value of the score based on the value of the activator
					# using the state name to value mapping (__name_to_value dictionary)
					if reg_element[-1]=='+':
						# this is a highest state activator
						# score is either zero or the max state
						if reg_element[0]=='!':
							y_sum += [int(N) if self.__name_to_value[reg_element[1:-1]]==N else 0]
						else:
							y_sum += [0 if self.__name_to_value[reg_element[:-1]]==N else int(N)]
					elif reg_element[0]=='!':
						# TODO: choose NOT function 
						y_sum += [int(self.discrete_not(self.__name_to_value[reg_element[1:]],N))]
					else:
						y_sum += [int(self.__name_to_value[reg_element])]

			if layer==0:
				# check for initializers and value of intializers
				if (self.__name_to_value[self.__regulated]==0 
					and len(y_init)!=0 
					and all([y==0 for y in y_init])==True):
					return 0
				else:
					# at the top layer, discrete OR (max) the groups that were split at commas
					# TODO: choose OR function 
					return sum(y_sum)

			else:
				return y_sum

	##############################################################
	## Functions for exporting truth tables and model rules
	##############################################################
	# these functions create boolean variables for elements

	def evaluate_state(self, state):
		""" determine the value of the regulated element
			for a given state (specific values of the inhibitors and activators),
			for now, this ignores delays
		"""

		self.__name_to_value.clear()

		for index in range(len(state)):
			# create a mapping between state names and state values
			self.__name_to_value[self.__name_list[index]] = int(state[index])

		# calculate activation and inhibition scores
		y_act = self.eval_reg(self.__act,0)
		y_inh = self.eval_reg(self.__inh,0)

		# define current element value, max states, and max delays for code readability
		X_curr = self.__name_to_value[self.__regulated]
		N = self.__max_state-1

		# Ignoring delays for now, but need to know spontaneous behavior
		# TODO: include delays in truth tables
		D_act = self.__spont_act
		D_inh = self.__spont_inh
		D = self.__delays

		# determine next value of the regulated element,
		# based on the type of regulators and activation/inhibition scores
		# TODO: some of this code could be condensed by reintroducing the gradient value
		# TODO: use difference between y_act and y_inh for strength of increase/decrease in next-state value
		if (self.__act) and (not self.__inh): 
			# this element has only positive regulators
			# set the next value to increase if activation > 0,
			# but don't increase beyond the maximum state.
			# if there is no activation, the value should spontaneously decay, 
			# but not below 0
			if (y_act > 0) and (X_curr < N):
				X_next = X_curr + 1
				self.__curr_delays[int(X_curr)] = 0
			elif (y_act > 0) and (X_curr == N):
				# at max state, hold value
				X_next = X_curr
			elif (y_act == 0) and (X_curr > 0):
				if D_inh != '':
					# spontaneously decay (inhibition)
					X_next = X_curr - 1
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_act == 0) and (X_curr == 0):
				X_next = X_curr
		elif (not self.__act) and (self.__inh):
			# this element has only negative regulators
			# set the next value to decrease if inhibition > 0,
			# but don't decrease below 0.
			# if there is no inhibition, the value should spontaneously increase, 
			# but not beyond the maximum state.
			if (y_inh > 0) and (X_curr > 0):
				X_next = X_curr - 1
			elif (y_inh > 0) and (X_curr == 0):
				X_next = X_curr
			elif (y_inh == 0) and (X_curr < N):
				if D_act != '':
					# spontaneously increase
					X_next = X_curr + 1
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_inh == 0) and (X_curr == N):
				X_next = X_curr		
		elif (self.__act) and (self.__inh):
			# this element has both activators and inhibitors
			# increase the value if activation > inhibition,
			# decrease if activation <= inhibition,
			# but not above/below the bounds of the max/min states.
			# Note: if activation == inhibition, 
			# the value spontaneously decays.
			if 	(y_act > y_inh) and (X_curr < N):
				X_next = X_curr + 1
			elif (y_act > y_inh) and (X_curr == N):
				X_next = X_curr
			elif (y_act == y_inh) and (X_curr > 0):
				if D_inh != '':
					# spontaneously decay (inhibition)
					X_next = X_curr - 1
				else:
					# no spontaneous behavior, hold value
					X_next = X_curr
			elif (y_act < y_inh) and (X_curr > 0):
				X_next = X_curr - 1
			elif (y_act <= y_inh) and (X_curr == 0):
				X_next = X_curr
		else:
			# this element has no regulators; 
			# Note that this shouldn't happen with the current model initialization
			X_next = X_curr
		
		# return the next state,
		# with a redundant sort to keep X_next within the state value bounds
		return sorted([0, X_next , N])[1]

	def generate_all_input_state(self,include_regulated=0): 
		# generate and record all possible input states 
		# basically generating a Boolean truth table, using all possible combinations of regulator values
		# the number of input states is given by max_state^(number of regulators)

		# TODO: need to include delays as variables in truth tables

		# get the total number of regulators
		# include_regulated input determines whether or not include the regulated element itself
		
		length = len(self.__name_list) if include_regulated else len(self.__name_list)-1
		total_states = []
		for num in range(int(math.pow(self.__max_state,length))):
			# generate the state
			result = ''
			temp = num
			while temp > 0:
				result = str(temp%self.__max_state) + result
				temp = temp//self.__max_state # integer division
			# record this state, padding with zeros on the left
			total_states.append(result.zfill(length))
		return total_states

	def generate_model_expression(self,output_model_file):
		if self.__act=='' and self.__inh=='':
			return None

		else:
			# generate truth table
			input_states = self.generate_all_input_state(1)


			bit_length = int(math.ceil(math.log(self.__max_state,2)))
			mode_to_expresion = [[] for x in range(bit_length)]

			# define the value for each state 
			for state in input_states:
				value = self.evaluate_state(state)
				for k in range(math.ceil(math.log(value+1,2))):
					if value%2:
						mode_to_expresion[k].append('('+self.state_to_expression(state)+')')
					value = value//2

			# write the model to a txt file
			output_model_file.write('{\n')

			# only use the underscore bit# notation if there is more than one bit needed for this element's value
			if bit_length > 1:
				for index in range(bit_length):
					mode = mode_to_expresion[index]
					if len(mode)!=0:
						output_model_file.write(self.__regulated+'_'+str(index)+' = '+\
							' + '.join(mode)+';\n')
					else:
						output_model_file.write(self.__regulated+'_'+str(index)+' = Const_False;\n')
			else:
				mode = mode_to_expresion[0]
				if len(mode)!=0:
					output_model_file.write(self.__regulated+' = '+\
						' + '.join(mode)+';\n')
				else:
					output_model_file.write(self.__regulated+' = Const_False;\n')

			output_model_file.write('}\n')

	def state_to_expression(self,state):
		# create a logical expression for the state
		# in sum of products form
		# TODO: use something like sympy to simplify the Boolean functions before writing to the output file
		result = list()
		
		for index in range(len(state)):
			element = self.__name_list[index]
			value = int(state[index])

			bit_length = int(math.ceil(math.log(self.__max_state,2)))

			# only use underscore bit# notation if there is more than one bit needed for this element's value
			if bit_length > 1:
				for k in range(bit_length):
					if value%2:
						result.append(element+'_'+str(k))
					else:
						result.append('!'+element+'_'+str(k))
					value = value//2
			else:
				if value%2:
					result.append(element)
				else:
					result.append('!'+element)

		return '*'.join(result)


	##############################################################
	## Utility functions
	##############################################################

	def discrete_not(self, x, N):
		"""
		"""
		if N >= x:
			return (N - x)
		else:
			raise ValueError('Can''t compute NOT, input is greater than max state')


	def split_comma_outside_parentheses(self, sentence):
		final_list = list()
		parentheses = 0
		start = 0
		for index in range(len(sentence)):
			char = sentence[index]
			if index==len(sentence)-1:
				final_list.append(sentence[start:index+1])
			elif char=='(' or char=='{' or char=='[':
				parentheses += 1
			elif char==')' or char=='}' or char==']':
				parentheses -= 1
			elif (char==',' and parentheses==0):
				final_list.append(sentence[start:index])
				start = index+1
		return final_list
